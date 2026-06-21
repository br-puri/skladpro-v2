import psycopg2
import psycopg2.extras
import json
import io
import os
import smtplib
import ssl
import subprocess
import tempfile
import time
import uuid
from email.message import EmailMessage
from datetime import date, datetime
from functools import wraps
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'skladpro-secret-2026')
DATABASE_URL = os.environ.get('DATABASE_URL', '')
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads', 'products')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp', 'gif', 'avif'}
CHROME_PATH = os.environ.get('CHROME_PATH', '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome')

# One-time tokens for headless catalog rendering (no auth needed for the renderer)
_catalog_tokens: dict = {}

PUBLIC_ROUTES = {'login', 'static', 'catalog_render'}

@app.before_request
def require_login():
    if request.endpoint and request.endpoint not in PUBLIC_ROUTES:
        if not session.get('user_id'):
            return redirect(url_for('login', next=request.path))


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Admin access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


@app.after_request
def write_audit(response):
    if (request.method != 'GET'
            and session.get('user_id')
            and request.endpoint not in (None, 'static', 'logout')
            and response.status_code < 500):
        try:
            with get_db() as db:
                db.execute(
                    "INSERT INTO audit_log(ts,user_id,username,method,path,status) VALUES(%s,%s,%s,%s,%s,%s)",
                    (datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'),
                     session['user_id'], session.get('username'),
                     request.method, request.path, response.status_code))
                db.commit()
        except Exception:
            pass
    return response


@app.context_processor
def inject_defaults():
    with get_db() as db:
        wh = db.execute("SELECT id FROM warehouses WHERE LOWER(name) LIKE %s LIMIT 1", ('%brent%',)).fetchone()
    return {'default_wh_id': wh['id'] if wh else None}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ── DB helpers ───────────────────────────────────────────────────────────────

class _Db:
    def __init__(self):
        self._conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    def __enter__(self):
        return self
    def __exit__(self, exc, *_):
        if exc:
            self._conn.rollback()
        else:
            self._conn.commit()
        self._conn.close()
    def execute(self, sql, params=None):
        cur = self._conn.cursor()
        cur.execute(sql, params or ())
        return cur
    def commit(self):
        self._conn.commit()
    def rollback(self):
        self._conn.rollback()


def get_db():
    return _Db()


def init_db():
    with get_db() as db:
        db.execute("""CREATE TABLE IF NOT EXISTS warehouses (
            id       SERIAL PRIMARY KEY,
            name     TEXT NOT NULL,
            location TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS products (
            id          SERIAL PRIMARY KEY,
            sku         TEXT DEFAULT '',
            barcode     TEXT DEFAULT '',
            name        TEXT NOT NULL,
            category    TEXT DEFAULT '',
            unit        TEXT DEFAULT 'pcs',
            cost        REAL DEFAULT 0,
            price       REAL DEFAULT 0,
            min_stock   REAL DEFAULT 0,
            description TEXT DEFAULT '',
            in_catalog  INTEGER DEFAULT 1
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS catalog_category_order (
            category   TEXT PRIMARY KEY,
            sort_order INTEGER DEFAULT 0
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS catalog_subcategory_order (
            category    TEXT,
            subcategory TEXT,
            sort_order  INTEGER DEFAULT 0,
            PRIMARY KEY (category, subcategory)
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS catalog_product_order (
            product_id INTEGER PRIMARY KEY,
            sort_order INTEGER DEFAULT 0
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS stock (
            product_id   INTEGER REFERENCES products(id) ON DELETE CASCADE,
            warehouse_id INTEGER REFERENCES warehouses(id) ON DELETE CASCADE,
            qty          REAL DEFAULT 0,
            PRIMARY KEY (product_id, warehouse_id)
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS contacts (
            id           SERIAL PRIMARY KEY,
            name         TEXT NOT NULL,
            type         TEXT DEFAULT 'customer',
            company      TEXT DEFAULT '',
            vat_number   TEXT DEFAULT '',
            phone        TEXT DEFAULT '',
            email        TEXT DEFAULT '',
            address      TEXT DEFAULT '',
            city         TEXT DEFAULT '',
            postcode     TEXT DEFAULT '',
            country      TEXT DEFAULT '',
            balance      REAL DEFAULT 0,
            discount_pct REAL DEFAULT 0,
            currency     TEXT DEFAULT 'GBP',
            notes        TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS contact_history (
            id         SERIAL PRIMARY KEY,
            contact_id INTEGER REFERENCES contacts(id) ON DELETE CASCADE,
            event_date TEXT NOT NULL,
            event_type TEXT DEFAULT 'note',
            "desc"     TEXT NOT NULL
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS purchases (
            id           SERIAL PRIMARY KEY,
            num          TEXT NOT NULL,
            doc_date     TEXT NOT NULL,
            supplier     TEXT NOT NULL,
            supplier_id  INTEGER REFERENCES contacts(id),
            warehouse_id INTEGER REFERENCES warehouses(id),
            subtotal     REAL DEFAULT 0,
            total        REAL DEFAULT 0,
            tax_pct      REAL DEFAULT 0,
            tax_amount   REAL DEFAULT 0,
            currency     TEXT DEFAULT 'GBP',
            status       TEXT DEFAULT 'received',
            notes        TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS purchase_items (
            id          SERIAL PRIMARY KEY,
            purchase_id INTEGER REFERENCES purchases(id) ON DELETE CASCADE,
            product_id  INTEGER REFERENCES products(id),
            qty         REAL DEFAULT 0,
            price       REAL DEFAULT 0
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS sales (
            id          SERIAL PRIMARY KEY,
            num         TEXT NOT NULL,
            doc_date    TEXT NOT NULL,
            customer    TEXT NOT NULL,
            customer_id INTEGER REFERENCES contacts(id),
            subtotal    REAL DEFAULT 0,
            total       REAL DEFAULT 0,
            discount    REAL DEFAULT 0,
            tax_pct     REAL DEFAULT 0,
            tax_amount  REAL DEFAULT 0,
            currency    TEXT DEFAULT 'GBP',
            status      TEXT DEFAULT 'pending',
            notes       TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS sale_items (
            id           SERIAL PRIMARY KEY,
            sale_id      INTEGER REFERENCES sales(id) ON DELETE CASCADE,
            product_id   INTEGER REFERENCES products(id),
            warehouse_id INTEGER REFERENCES warehouses(id),
            qty          REAL DEFAULT 0,
            price        REAL DEFAULT 0,
            discount_pct REAL DEFAULT 0
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS transfers (
            id           SERIAL PRIMARY KEY,
            doc_date     TEXT NOT NULL,
            from_wh_id   INTEGER REFERENCES warehouses(id),
            to_wh_id     INTEGER REFERENCES warehouses(id),
            product_id   INTEGER REFERENCES products(id),
            qty          REAL DEFAULT 0,
            notes        TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS writeoffs (
            id           SERIAL PRIMARY KEY,
            num          TEXT NOT NULL,
            doc_date     TEXT NOT NULL,
            warehouse_id INTEGER REFERENCES warehouses(id),
            reason       TEXT DEFAULT '',
            total_cost   REAL DEFAULT 0,
            notes        TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS writeoff_items (
            id          SERIAL PRIMARY KEY,
            writeoff_id INTEGER REFERENCES writeoffs(id) ON DELETE CASCADE,
            product_id  INTEGER REFERENCES products(id),
            qty         REAL DEFAULT 0,
            cost        REAL DEFAULT 0
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS inventory_counts (
            id           SERIAL PRIMARY KEY,
            num          TEXT NOT NULL,
            doc_date     TEXT NOT NULL,
            warehouse_id INTEGER REFERENCES warehouses(id),
            status       TEXT DEFAULT 'draft',
            notes        TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS inventory_count_items (
            id         SERIAL PRIMARY KEY,
            count_id   INTEGER REFERENCES inventory_counts(id) ON DELETE CASCADE,
            product_id INTEGER REFERENCES products(id),
            qty_system REAL DEFAULT 0,
            qty_actual REAL DEFAULT 0
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS transactions (
            id         SERIAL PRIMARY KEY,
            doc_date   TEXT NOT NULL,
            type       TEXT DEFAULT 'income',
            method     TEXT DEFAULT 'cash',
            "desc"     TEXT NOT NULL,
            amount     REAL DEFAULT 0,
            currency   TEXT DEFAULT 'GBP',
            contact_id INTEGER REFERENCES contacts(id),
            contact    TEXT DEFAULT '',
            ref_doc    TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS exchange_rates (
            id         SERIAL PRIMARY KEY,
            currency   TEXT NOT NULL,
            rate       REAL NOT NULL,
            set_date   TEXT NOT NULL,
            fetched_at TEXT
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS ttns (
            id        SERIAL PRIMARY KEY,
            num       TEXT NOT NULL,
            doc_date  TEXT NOT NULL,
            sender    TEXT DEFAULT '',
            receiver  TEXT DEFAULT '',
            carrier   TEXT DEFAULT '',
            ref_doc   TEXT DEFAULT '',
            status    TEXT DEFAULT 'in_transit',
            notes     TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS discounts (
            id         SERIAL PRIMARY KEY,
            name       TEXT NOT NULL,
            type       TEXT DEFAULT 'percent',
            value      REAL DEFAULT 0,
            applies_to TEXT DEFAULT 'all',
            active     INTEGER DEFAULT 1
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS categories (
            id          SERIAL PRIMARY KEY,
            name        TEXT NOT NULL UNIQUE,
            description TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS quotes (
            id          SERIAL PRIMARY KEY,
            num         TEXT NOT NULL,
            doc_date    TEXT NOT NULL,
            expiry_date TEXT DEFAULT '',
            customer    TEXT NOT NULL,
            customer_id INTEGER REFERENCES contacts(id),
            subtotal    REAL DEFAULT 0,
            total       REAL DEFAULT 0,
            discount    REAL DEFAULT 0,
            tax_pct     REAL DEFAULT 0,
            tax_amount  REAL DEFAULT 0,
            currency    TEXT DEFAULT 'GBP',
            status      TEXT DEFAULT 'draft',
            notes       TEXT DEFAULT ''
        )""")
        db.execute("""CREATE TABLE IF NOT EXISTS quote_items (
            id           SERIAL PRIMARY KEY,
            quote_id     INTEGER REFERENCES quotes(id) ON DELETE CASCADE,
            product_id   INTEGER REFERENCES products(id),
            warehouse_id INTEGER REFERENCES warehouses(id),
            qty          REAL DEFAULT 0,
            price        REAL DEFAULT 0,
            discount_pct REAL DEFAULT 0
        )""")
        for col, dflt in [('company','""'), ('vat_number','""'), ('address2','""'), ('city','""'), ('postcode','""'), ('country','""')]:
            try:
                db.execute(f"ALTER TABLE contacts ADD COLUMN IF NOT EXISTS {col} TEXT DEFAULT ''")
            except Exception:
                db.rollback()
        try:
            db.execute("ALTER TABLE sales ADD COLUMN IF NOT EXISTS archived INTEGER DEFAULT 0")
        except Exception:
            db.rollback()
        try:
            db.execute("ALTER TABLE sales ADD COLUMN IF NOT EXISTS paid INTEGER DEFAULT 0")
        except Exception:
            db.rollback()
        try:
            db.execute("ALTER TABLE sale_items ADD COLUMN IF NOT EXISTS product_name_snap TEXT DEFAULT ''")
        except Exception:
            db.rollback()
        for col in ["photo TEXT DEFAULT ''", 'length REAL DEFAULT 0', 'width REAL DEFAULT 0',
                    'height REAL DEFAULT 0', 'weight REAL DEFAULT 0', 'cbm REAL DEFAULT 0',
                    'carton_qty REAL DEFAULT 0', "subcategory TEXT DEFAULT ''",
                    'ctn_price REAL DEFAULT 0',
                    "china_price REAL DEFAULT 0", "china_currency TEXT DEFAULT 'RMB'"]:
            try:
                db.execute(f"ALTER TABLE products ADD COLUMN IF NOT EXISTS {col}")
            except Exception:
                db.rollback()
        for col in ['discount REAL DEFAULT 0']:
            try:
                db.execute(f"ALTER TABLE purchases ADD COLUMN IF NOT EXISTS {col}")
            except Exception:
                db.rollback()
        try:
            db.execute("ALTER TABLE categories ADD COLUMN IF NOT EXISTS parent_id INTEGER DEFAULT NULL")
        except Exception:
            db.rollback()
        db.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id       SERIAL PRIMARY KEY,
                username TEXT NOT NULL UNIQUE,
                password TEXT NOT NULL,
                role     TEXT DEFAULT 'admin'
            )
        """)
        db.commit()
        db.execute("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id         SERIAL PRIMARY KEY,
                ts         TEXT NOT NULL,
                user_id    INTEGER,
                username   TEXT,
                method     TEXT,
                path       TEXT,
                status     INTEGER
            )
        """)
        db.execute("""CREATE TABLE IF NOT EXISTS credit_notes (
                id          SERIAL PRIMARY KEY,
                num         TEXT NOT NULL,
                doc_date    TEXT NOT NULL,
                sale_id     INTEGER REFERENCES sales(id),
                customer    TEXT NOT NULL,
                customer_id INTEGER REFERENCES contacts(id),
                warehouse_id INTEGER REFERENCES warehouses(id),
                total       REAL DEFAULT 0,
                status      TEXT DEFAULT 'draft',
                notes       TEXT DEFAULT '',
                created_at  TEXT
            )""")
        db.execute("""CREATE TABLE IF NOT EXISTS credit_note_items (
                id             SERIAL PRIMARY KEY,
                credit_note_id INTEGER NOT NULL REFERENCES credit_notes(id),
                product_id     INTEGER REFERENCES products(id),
                qty            REAL NOT NULL,
                price          REAL NOT NULL,
                product_name   TEXT DEFAULT ''
            )""")
        db.execute("""CREATE TABLE IF NOT EXISTS debit_notes (
                id           SERIAL PRIMARY KEY,
                num          TEXT NOT NULL,
                doc_date     TEXT NOT NULL,
                purchase_id  INTEGER REFERENCES purchases(id),
                supplier     TEXT NOT NULL,
                supplier_id  INTEGER REFERENCES contacts(id),
                warehouse_id INTEGER REFERENCES warehouses(id),
                total        REAL DEFAULT 0,
                status       TEXT DEFAULT 'draft',
                notes        TEXT DEFAULT '',
                created_at   TEXT
            )""")
        db.execute("""CREATE TABLE IF NOT EXISTS debit_note_items (
                id            SERIAL PRIMARY KEY,
                debit_note_id INTEGER NOT NULL REFERENCES debit_notes(id),
                product_id    INTEGER REFERENCES products(id),
                qty           REAL NOT NULL,
                price         REAL NOT NULL,
                product_name  TEXT DEFAULT ''
            )""")
        db.execute("""CREATE TABLE IF NOT EXISTS supplier_orders (
                id          SERIAL PRIMARY KEY,
                num         TEXT NOT NULL,
                doc_date    TEXT NOT NULL,
                supplier    TEXT DEFAULT '',
                supplier_id INTEGER REFERENCES contacts(id),
                notes       TEXT DEFAULT '',
                status      TEXT DEFAULT 'draft'
            )""")
        db.execute("""CREATE TABLE IF NOT EXISTS supplier_order_items (
                id           SERIAL PRIMARY KEY,
                order_id     INTEGER NOT NULL REFERENCES supplier_orders(id) ON DELETE CASCADE,
                name         TEXT DEFAULT '',
                article      TEXT DEFAULT '',
                photo        TEXT DEFAULT '',
                cbm          REAL DEFAULT 0,
                ctns         REAL DEFAULT 0,
                pcs_per_ctn  REAL DEFAULT 0
            )""")
        db.execute("""CREATE TABLE IF NOT EXISTS invoice_payments (
                id           SERIAL PRIMARY KEY,
                sale_id      INTEGER NOT NULL REFERENCES sales(id),
                payment_date TEXT NOT NULL,
                amount       REAL NOT NULL,
                method       TEXT DEFAULT 'bank',
                notes        TEXT DEFAULT '',
                created_at   TEXT
            )""")
        db.commit()
        # Migrations — safe to run on existing DBs
        try:
            db.execute("ALTER TABLE exchange_rates ADD COLUMN IF NOT EXISTS fetched_at TEXT")
            db.commit()
        except Exception:
            db.rollback()
        # Create default admin if no users exist
        if db.execute("SELECT COUNT(*) AS n FROM users").fetchone()['n'] == 0:
            db.execute("INSERT INTO users(username,password,role) VALUES(%s,%s,%s)",
                       ('admin', generate_password_hash('admin123', method='pbkdf2:sha256'), 'admin'))
            db.commit()
        if db.execute("SELECT COUNT(*) AS n FROM warehouses").fetchone()['n'] == 0:
            db.execute("INSERT INTO warehouses(name,location) VALUES('Main Warehouse','Kyiv, Boryspilska 14')")
            db.execute("INSERT INTO warehouses(name,location) VALUES('South Depot','Odesa, Morska 3')")
            db.execute("INSERT INTO products(sku,barcode,name,category,unit,cost,price,min_stock) VALUES('P001','4820000000011','Steel Pipe 50mm','Pipes','pcs',85,120,5)")
            db.execute("INSERT INTO products(sku,barcode,name,category,unit,cost,price,min_stock) VALUES('P002','4820000000022','Copper Wire 2.5mm','Electrical','m',11,18,20)")
            db.execute("INSERT INTO products(sku,barcode,name,category,unit,cost,price,min_stock) VALUES('P003','4820000000033','Cement 25kg','Building','bag',22,35,10)")
            db.execute("INSERT INTO products(sku,barcode,name,category,unit,cost,price,min_stock) VALUES('P004','4820000000044','PVC Sheet 4x8','Plastics','sheet',140,210,3)")
            for pid, wid, qty in [(1,1,45),(1,2,12),(2,1,320),(3,1,5),(3,2,80),(4,2,28)]:
                db.execute("INSERT INTO stock VALUES(%s,%s,%s) ON CONFLICT (product_id, warehouse_id) DO UPDATE SET qty=EXCLUDED.qty", (pid, wid, qty))
            db.execute("INSERT INTO contacts(name,type,phone,email,address,balance,discount_pct) VALUES('Alfa Supplies LLC','supplier','+380 44 123 4567','orders@alfa.ua','Kyiv, Promyslova 5',-4200,0)")
            db.execute("INSERT INTO contacts(name,type,phone,email,address,balance,discount_pct) VALUES('Build Masters Co','customer','+380 67 987 6543','buy@build.ua','Kyiv, Budivelnykiv 12',1800,5)")
            db.execute("INSERT INTO contacts(name,type,phone,email,address,balance,discount_pct) VALUES('TechElec Group','customer','+380 50 555 0099','info@techelec.ua','Odesa, Elektrychna 3',0,0)")
            db.execute('INSERT INTO transactions(doc_date,type,method,"desc",amount,contact) VALUES(\'2026-03-12\',\'income\',\'cash\',\'Payment SAL-001\',1500,\'Build Masters Co\')')
            db.execute('INSERT INTO transactions(doc_date,type,method,"desc",amount,contact) VALUES(\'2026-03-10\',\'expense\',\'bank\',\'Payment PUR-001\',1700,\'Alfa Supplies LLC\')')
            db.execute("INSERT INTO exchange_rates(currency,rate,set_date) VALUES('USD',41.5,'2026-03-01')")
            db.execute("INSERT INTO exchange_rates(currency,rate,set_date) VALUES('EUR',44.8,'2026-03-01')")
            db.execute("INSERT INTO discounts(name,type,value,applies_to) VALUES('Summer Sale','percent',10,'all')")
            db.execute("INSERT INTO discounts(name,type,value,applies_to) VALUES('VIP Customer','percent',15,'all')")
            db.commit()


def next_num(prefix, table):
    ym = date.today().strftime('%Y%m')
    with get_db() as db:
        n = db.execute(f"SELECT COUNT(*) AS n FROM {table} WHERE num LIKE %s", (f'{prefix}-{ym}%',)).fetchone()['n'] + 1
    return f"{prefix}-{ym}-{n:03d}"


def get_settings():
    with get_db() as db:
        rows = db.execute("SELECT key, value FROM settings").fetchall()
    return {r['key']: r['value'] for r in rows}


@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('user_id'):
        return redirect(url_for('dashboard'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        with get_db() as db:
            user = db.execute("SELECT * FROM users WHERE username=%s", (username,)).fetchone()
        if user and check_password_hash(user['password'], password):
            session.permanent = True
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            next_page = request.args.get('next') or url_for('dashboard')
            return redirect(next_page)
        error = 'Invalid username or password'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── My Account ────────────────────────────────────────────────────────────────

@app.route('/account', methods=['GET', 'POST'])
def my_account():
    if request.method == 'POST':
        current_pw = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')
        if not new_pw:
            flash('New password cannot be empty', 'error')
            return redirect(url_for('my_account'))
        if new_pw != confirm_pw:
            flash('New passwords do not match', 'error')
            return redirect(url_for('my_account'))
        with get_db() as db:
            user = db.execute("SELECT * FROM users WHERE id=%s", (session['user_id'],)).fetchone()
            if not check_password_hash(user['password'], current_pw):
                flash('Current password is incorrect', 'error')
                return redirect(url_for('my_account'))
            db.execute("UPDATE users SET password=%s WHERE id=%s",
                       (generate_password_hash(new_pw, method='pbkdf2:sha256'), session['user_id']))
            db.commit()
        flash('Password changed successfully', 'success')
        return redirect(url_for('my_account'))
    return render_template('my_account.html')


# ── User Management ───────────────────────────────────────────────────────────

@app.route('/users')
@admin_required
def users_list():
    with get_db() as db:
        users = db.execute("SELECT id, username, role FROM users ORDER BY username").fetchall()
    return render_template('users.html', users=users)


@app.route('/users/add', methods=['POST'])
@admin_required
def add_user():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    role = request.form.get('role', 'viewer')
    if not username or not password:
        flash('Username and password are required', 'error')
        return redirect(url_for('users_list'))
    if role not in ('admin', 'viewer'):
        role = 'viewer'
    with get_db() as db:
        try:
            db.execute("INSERT INTO users(username,password,role) VALUES(%s,%s,%s)",
                       (username, generate_password_hash(password, method='pbkdf2:sha256'), role))
            db.commit()
            flash(f'User "{username}" created', 'success')
        except Exception:
            flash('Username already exists', 'error')
    return redirect(url_for('users_list'))


@app.route('/users/<int:uid>/delete', methods=['POST'])
@admin_required
def delete_user(uid):
    if uid == session.get('user_id'):
        flash('Cannot delete your own account', 'error')
        return redirect(url_for('users_list'))
    with get_db() as db:
        admin_count = db.execute("SELECT COUNT(*) AS n FROM users WHERE role='admin'").fetchone()['n']
        user = db.execute("SELECT * FROM users WHERE id=%s", (uid,)).fetchone()
        if user and user['role'] == 'admin' and admin_count <= 1:
            flash('Cannot delete the last admin account', 'error')
            return redirect(url_for('users_list'))
        db.execute("DELETE FROM users WHERE id=%s", (uid,))
        db.commit()
    flash('User deleted', 'success')
    return redirect(url_for('users_list'))


@app.route('/users/<int:uid>/reset-password', methods=['POST'])
@admin_required
def reset_user_password(uid):
    password = request.form.get('password', '')
    if not password:
        flash('Password cannot be empty', 'error')
        return redirect(url_for('users_list'))
    with get_db() as db:
        db.execute("UPDATE users SET password=%s WHERE id=%s",
                   (generate_password_hash(password, method='pbkdf2:sha256'), uid))
        db.commit()
    flash('Password updated', 'success')
    return redirect(url_for('users_list'))


@app.route('/users/<int:uid>/role', methods=['POST'])
@admin_required
def change_user_role(uid):
    role = request.form.get('role', 'viewer')
    if role not in ('admin', 'viewer'):
        role = 'viewer'
    if uid == session.get('user_id') and role != 'admin':
        flash('Cannot remove admin role from your own account', 'error')
        return redirect(url_for('users_list'))
    with get_db() as db:
        if role != 'admin':
            admin_count = db.execute("SELECT COUNT(*) AS n FROM users WHERE role='admin'").fetchone()['n']
            user = db.execute("SELECT role FROM users WHERE id=%s", (uid,)).fetchone()
            if user and user['role'] == 'admin' and admin_count <= 1:
                flash('Cannot remove the last admin', 'error')
                return redirect(url_for('users_list'))
        db.execute("UPDATE users SET role=%s WHERE id=%s", (role, uid))
        db.commit()
    flash('Role updated', 'success')
    return redirect(url_for('users_list'))


@app.route('/audit-log')
@admin_required
def audit_log():
    page = int(request.args.get('page', 1))
    per_page = 100
    offset = (page - 1) * per_page
    user_filter = request.args.get('user', '')
    with get_db() as db:
        base = "FROM audit_log WHERE 1=1"
        params = []
        if user_filter:
            base += " AND username=%s"
            params.append(user_filter)
        total = db.execute(f"SELECT COUNT(*) AS n {base}", params).fetchone()['n']
        rows = db.execute(
            f"SELECT * {base} ORDER BY id DESC LIMIT %s OFFSET %s",
            params + [per_page, offset]).fetchall()
        users = db.execute("SELECT DISTINCT username FROM audit_log ORDER BY username").fetchall()
    pages = (total + per_page - 1) // per_page
    return render_template('audit_log.html', rows=rows, page=page, pages=pages,
                           total=total, user_filter=user_filter, users=users)


@app.route('/settings', methods=['GET', 'POST'])
@admin_required
def settings_page():
    if request.method == 'POST':
        fields = ['co_name', 'co_company', 'co_address', 'co_address2', 'co_city', 'co_postcode',
                  'co_country', 'co_vat', 'co_email', 'co_phone',
                  'co_delivery_address', 'co_delivery_address2', 'co_delivery_city', 'co_delivery_postcode', 'co_delivery_country',
                  'co_bank_name', 'co_sort_code', 'co_account_number',
                  'smtp_host', 'smtp_port', 'smtp_user', 'smtp_pass', 'smtp_from', 'smtp_tls']
        # Boolean toggles (checkboxes)
        with get_db() as db:
            db.execute("INSERT INTO settings(key,value) VALUES(%s,%s) ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value",
                       ('co_hide_stock', '1' if request.form.get('co_hide_stock') else ''))
        with get_db() as db:
            for f in fields:
                db.execute("INSERT INTO settings(key,value) VALUES(%s,%s) ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value",
                           (f, request.form.get(f, '')))
            logo_file = request.files.get('co_logo')
            if logo_file and logo_file.filename and allowed_file(logo_file.filename):
                ext = logo_file.filename.rsplit('.', 1)[1].lower()
                logo_path = os.path.join(UPLOAD_FOLDER, '..', f'logo.{ext}')
                logo_path = os.path.normpath(logo_path)
                logo_file.save(logo_path)
                db.execute("INSERT INTO settings(key,value) VALUES(%s,%s) ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value",
                           ('co_logo', f'logo.{ext}'))
            db.commit()
        flash('Settings saved', 'success')
        return redirect(url_for('settings_page'))
    s = get_settings()
    return render_template('settings.html', s=s)


# ── Dashboard ────────────────────────────────────────────────────────────────

@app.route('/')
def dashboard():
    with get_db() as db:
        products = db.execute("SELECT * FROM products").fetchall()
        warehouses = db.execute("SELECT * FROM warehouses").fetchall()
        sales = db.execute("SELECT * FROM sales ORDER BY doc_date DESC LIMIT 8").fetchall()
        transactions = db.execute("SELECT * FROM transactions").fetchall()
        stock_value = sum(
            db.execute("SELECT COALESCE(SUM(qty),0) AS s FROM stock WHERE product_id=%s", (p['id'],)).fetchone()['s'] * p['cost']
            for p in products
        )
        completed_revenue = db.execute("SELECT COALESCE(SUM(total),0) AS s FROM sales WHERE status='completed'").fetchone()['s']
        pending_sales = db.execute("SELECT COUNT(*) AS n FROM sales WHERE status='pending'").fetchone()['n']
        low_stock = []
        for p in products:
            qty = db.execute("SELECT COALESCE(SUM(qty),0) AS s FROM stock WHERE product_id=%s", (p['id'],)).fetchone()['s']
            if qty <= p['min_stock']:
                low_stock.append({'name': p['name'], 'qty': qty, 'min_stock': p['min_stock'], 'unit': p['unit']})
        income = sum(t['amount'] for t in transactions if t['type'] == 'income')
        expenses = sum(t['amount'] for t in transactions if t['type'] == 'expense')
        monthly = db.execute("""SELECT to_char(doc_date::date, 'YYYY-MM') as month,
            SUM(total) as total, COUNT(*) as count
            FROM sales WHERE status='completed'
            GROUP BY month ORDER BY month DESC LIMIT 12""").fetchall()
        top_products = db.execute("""
            SELECT p.name, SUM(si.qty * si.price) as revenue, SUM(si.qty) as units
            FROM sale_items si
            JOIN products p ON p.id = si.product_id
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status = 'completed'
            GROUP BY si.product_id, p.name ORDER BY revenue DESC LIMIT 6
        """).fetchall()
        status_counts = db.execute("""
            SELECT status, COUNT(*) as cnt, COALESCE(SUM(total),0) as total
            FROM sales GROUP BY status
        """).fetchall()
    return render_template('dashboard.html',
        product_count=len(products), warehouse_count=len(warehouses),
        stock_value=stock_value, completed_revenue=completed_revenue,
        pending_sales=pending_sales, net_cash=income - expenses,
        recent_sales=sales, low_stock=low_stock,
        monthly_sales=list(reversed([dict(m) for m in monthly])),
        top_products=[dict(r) for r in top_products],
        status_counts=[dict(r) for r in status_counts])


# ── Products ─────────────────────────────────────────────────────────────────

@app.route('/products')
def products():
    q = request.args.get('q', '')
    cat = request.args.get('cat', '')
    subcat = request.args.get('subcat', '')
    with get_db() as db:
        base = "SELECT * FROM products WHERE 1=1"
        params = []
        if q:
            base += " AND (name LIKE %s OR sku LIKE %s OR barcode LIKE %s)"
            params += [f'%{q}%', f'%{q}%', f'%{q}%']
        if cat:
            base += " AND category=%s"
            params.append(cat)
        if subcat:
            base += " AND subcategory=%s"
            params.append(subcat)
        rows = db.execute(base + " ORDER BY name", params).fetchall()
        warehouses = db.execute("SELECT * FROM warehouses").fetchall()
        categories = [dict(r) for r in db.execute("SELECT DISTINCT category FROM products WHERE category!='' ORDER BY category").fetchall()]
        subcategories = [dict(r) for r in db.execute("SELECT DISTINCT category, subcategory FROM products WHERE subcategory!='' ORDER BY subcategory").fetchall()]
        result = []
        for p in rows:
            stock_by_wh, total = {}, 0
            for w in warehouses:
                r = db.execute("SELECT qty FROM stock WHERE product_id=%s AND warehouse_id=%s", (p['id'], w['id'])).fetchone()
                q_val = r['qty'] if r else 0
                stock_by_wh[w['id']] = q_val
                total += q_val
            result.append({'product': p, 'stock': stock_by_wh, 'total': total, 'low': total <= p['min_stock']})
    return render_template('products.html', products=result, warehouses=warehouses,
                           categories=categories, subcategories=subcategories, q=q, cat=cat, subcat=subcat)


@app.route('/products/catalog')
def product_catalog():
    ids = request.args.get('ids', '')
    with get_db() as db:
        if ids:
            id_list = [int(i) for i in ids.split(',') if i.strip().isdigit()]
            placeholders = ','.join(['%s'] * len(id_list))
            rows = db.execute(f"SELECT * FROM products WHERE id IN ({placeholders})", id_list).fetchall()
        else:
            rows = db.execute("SELECT * FROM products WHERE in_catalog=1").fetchall()
        # stock totals per product
        stock_rows = db.execute("SELECT product_id, SUM(qty) AS total FROM stock GROUP BY product_id").fetchall()
        stock_map = {r['product_id']: r['total'] for r in stock_rows}
        cat_order = {r['category']: r['sort_order'] for r in db.execute("SELECT category, sort_order FROM catalog_category_order").fetchall()}
        sub_order = {(r['category'], r['subcategory']): r['sort_order'] for r in db.execute("SELECT category, subcategory, sort_order FROM catalog_subcategory_order").fetchall()}
        prod_order = {r['product_id']: r['sort_order'] for r in db.execute("SELECT product_id, sort_order FROM catalog_product_order").fetchall()}
    # Build flat products list, sort by category/subcategory/product custom order
    products = [dict(p) for p in rows]
    for p in products:
        p['stock_total'] = stock_map.get(p['id'], 0) or 0
    products.sort(key=lambda p: (
        cat_order.get(p['category'] or 'Uncategorised', 9999),
        p['category'] or '',
        sub_order.get((p['category'] or 'Uncategorised', p['subcategory'] or ''), 9999),
        p['subcategory'] or '',
        prod_order.get(p['id'], 9999),
        p['name'] or '',
    ))
    all_cats = {(p['category'] or 'Uncategorised') for p in products}
    categories = sorted(all_cats, key=lambda c: (cat_order.get(c, 9999), c))
    return render_template('product_catalog.html',
                           products=products,
                           categories=categories,
                           co=get_settings(),
                           now=date.today())


# ── Headless PDF renderer (public, one-time token) ───────────────────────────
def _catalog_data():
    """Shared helper: return products, categories, co, now for catalog."""
    with get_db() as db:
        rows = db.execute("SELECT * FROM products WHERE in_catalog=1").fetchall()
        stock_rows = db.execute("SELECT product_id, SUM(qty) AS total FROM stock GROUP BY product_id").fetchall()
        cat_order = {r['category']: r['sort_order'] for r in db.execute("SELECT category, sort_order FROM catalog_category_order").fetchall()}
        sub_order = {(r['category'], r['subcategory']): r['sort_order'] for r in db.execute("SELECT category, subcategory, sort_order FROM catalog_subcategory_order").fetchall()}
        prod_order = {r['product_id']: r['sort_order'] for r in db.execute("SELECT product_id, sort_order FROM catalog_product_order").fetchall()}
    stock_map = {r['product_id']: r['total'] for r in stock_rows}
    products = [dict(p) for p in rows]
    for p in products:
        p['stock_total'] = stock_map.get(p['id'], 0) or 0
    products.sort(key=lambda p: (
        cat_order.get(p['category'] or 'Uncategorised', 9999),
        p['category'] or '',
        sub_order.get((p['category'] or 'Uncategorised', p['subcategory'] or ''), 9999),
        p['subcategory'] or '',
        prod_order.get(p['id'], 9999),
        p['name'] or '',
    ))
    all_cats = {(p['category'] or 'Uncategorised') for p in products}
    categories = sorted(all_cats, key=lambda c: (cat_order.get(c, 9999), c))
    return products, categories


@app.route('/products/catalog/render')
def catalog_render():
    """Token-gated, auth-free render used by Chrome headless for PDF export."""
    token = request.args.get('token', '')
    now = time.time()
    # Clean up expired tokens
    expired = [k for k, v in _catalog_tokens.items() if v < now]
    for k in expired:
        _catalog_tokens.pop(k, None)
    if token not in _catalog_tokens:
        return 'Forbidden', 403
    _catalog_tokens.pop(token, None)
    products, categories = _catalog_data()
    return render_template('product_catalog.html',
                           products=products,
                           categories=categories,
                           co=get_settings(),
                           now=date.today())


@app.route('/products/catalog/pdf')
def catalog_pdf_download():
    """Generate a PDF of the catalog using Chrome headless and serve it."""
    if not os.path.exists(CHROME_PATH):
        flash('Chrome not found — cannot generate PDF.', 'error')
        return redirect(url_for('product_catalog'))

    # Issue a one-time token (valid 90 s)
    token = str(uuid.uuid4())
    _catalog_tokens[token] = time.time() + 90

    port = request.environ.get('SERVER_PORT', 5001)
    render_url = f'http://127.0.0.1:{port}/products/catalog/render?token={token}'

    pdf_fd, pdf_path = tempfile.mkstemp(suffix='.pdf')
    os.close(pdf_fd)

    try:
        subprocess.run(
            [
                CHROME_PATH,
                '--headless',
                '--disable-gpu',
                '--no-sandbox',
                '--run-all-compositor-stages-before-draw',
                '--no-pdf-header-footer',
                '--window-size=794,1123',   # A4 at 96 dpi (210×297 mm)
                '--print-to-pdf-no-header',
                f'--print-to-pdf={pdf_path}',
                render_url,
            ],
            capture_output=True,
            timeout=60,
        )
        with open(pdf_path, 'rb') as f:
            data = f.read()
    finally:
        os.unlink(pdf_path)
        _catalog_tokens.pop(token, None)  # remove if unused

    fname = f"Product_Catalogue_{date.today().strftime('%Y%m%d')}.pdf"
    response = send_file(io.BytesIO(data), mimetype='application/pdf',
                         download_name=fname, as_attachment=False)
    response.headers['Content-Disposition'] = f'inline; filename="{fname}"'
    return response


@app.route('/products/catalog/manage', methods=['GET', 'POST'])
@admin_required
def catalog_manage():
    with get_db() as db:
        if request.method == 'POST':
            checked = set(request.form.getlist('in_catalog'))
            all_ids = [r['id'] for r in db.execute("SELECT id FROM products").fetchall()]
            for pid in all_ids:
                db.execute("UPDATE products SET in_catalog=%s WHERE id=%s",
                           (1 if str(pid) in checked else 0, pid))
            # Save category order
            cat_order_list = request.form.getlist('cat_order[]')
            db.execute("DELETE FROM catalog_category_order")
            for idx, cat in enumerate(cat_order_list):
                db.execute("INSERT INTO catalog_category_order (category, sort_order) VALUES (%s,%s)",
                           (cat, idx))
            # Save subcategory order — each value is "category|||subcategory"
            sub_order_list = request.form.getlist('sub_order[]')
            db.execute("DELETE FROM catalog_subcategory_order")
            for idx, key in enumerate(sub_order_list):
                if '|||' in key:
                    c, s = key.split('|||', 1)
                    db.execute("INSERT INTO catalog_subcategory_order (category, subcategory, sort_order) VALUES (%s,%s,%s)",
                               (c, s, idx))
            # Save product order
            prod_order_list = request.form.getlist('prod_order[]')
            db.execute("DELETE FROM catalog_product_order")
            for idx, pid in enumerate(prod_order_list):
                if pid.isdigit():
                    db.execute("INSERT INTO catalog_product_order (product_id, sort_order) VALUES (%s,%s)",
                               (int(pid), idx))
            db.commit()
            flash('Catalog updated.', 'success')
            return redirect(url_for('product_catalog'))
        rows = db.execute(
            "SELECT id, name, sku, barcode, photo, category, COALESCE(subcategory,'') AS subcategory, in_catalog "
            "FROM products"
        ).fetchall()
        cat_order_rows = db.execute("SELECT category, sort_order FROM catalog_category_order ORDER BY sort_order").fetchall()
        cat_order = [r['category'] for r in cat_order_rows]
        sub_order_rows = db.execute("SELECT category, subcategory, sort_order FROM catalog_subcategory_order ORDER BY sort_order").fetchall()
        sub_order = {(r['category'], r['subcategory']): r['sort_order'] for r in sub_order_rows}
        prod_order_rows = db.execute("SELECT product_id, sort_order FROM catalog_product_order ORDER BY sort_order").fetchall()
        prod_order = {r['product_id']: r['sort_order'] for r in prod_order_rows}
    products = [dict(r) for r in rows]
    # group by category → subcategory, with products sorted by custom order
    cat_groups = {}
    for p in products:
        cat = p['category'] or 'Uncategorised'
        sub = p['subcategory'] or ''
        cat_groups.setdefault(cat, {}).setdefault(sub, []).append(p)
    # Sort products inside each subcategory by saved order
    for cat, subs in cat_groups.items():
        for sub, prods in subs.items():
            prods.sort(key=lambda p: (prod_order.get(p['id'], 9999), p['name'] or ''))
    # Sort categories: saved order first, then alphabetically
    ordered_cats = cat_order + sorted(c for c in cat_groups if c not in cat_order)
    # Sort subcategories inside each category
    cat_groups_sorted = {}
    for c in ordered_cats:
        if c not in cat_groups:
            continue
        subs = cat_groups[c]
        ordered_subs = sorted(subs.keys(), key=lambda s: (sub_order.get((c, s), 9999), s))
        cat_groups_sorted[c] = {s: subs[s] for s in ordered_subs}
    return render_template('catalog_manage.html', cat_groups=cat_groups_sorted,
                           total=len(products),
                           selected=sum(1 for p in products if p['in_catalog']))


@app.route('/categories')
def categories():
    with get_db() as db:
        all_cats = [dict(r) for r in db.execute("SELECT * FROM categories ORDER BY name").fetchall()]
    parents = [c for c in all_cats if not c['parent_id']]
    children = {}
    for c in all_cats:
        if c['parent_id']:
            children.setdefault(c['parent_id'], []).append(c)
    return render_template('categories.html', parents=parents, children=children)


@app.route('/categories/add', methods=['POST'])
@admin_required
def add_category():
    name = request.form.get('name', '').strip()
    parent_id = request.form.get('parent_id') or None
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.args.get('ajax')
    if name:
        with get_db() as db:
            try:
                cur = db.execute("INSERT INTO categories(name,description,parent_id) VALUES(%s,%s,%s)",
                                 (name, request.form.get('description', ''), parent_id))
                db.commit()
                if is_ajax:
                    return jsonify({'ok': True, 'id': cur.lastrowid, 'name': name,
                                    'parent_id': int(parent_id) if parent_id else None})
                flash('Category added', 'success')
            except Exception:
                if is_ajax:
                    return jsonify({'ok': False, 'error': 'Category already exists'}), 400
                flash('Category already exists', 'error')
    if is_ajax:
        return jsonify({'ok': False, 'error': 'Name required'}), 400
    return redirect(url_for('categories'))


@app.route('/categories/rename/<int:cid>', methods=['POST'])
@admin_required
def rename_category(cid):
    new_name = request.form.get('name', '').strip()
    new_desc = request.form.get('description', '').strip()
    if not new_name:
        flash('Name cannot be empty', 'error')
        return redirect(url_for('categories'))
    with get_db() as db:
        row = db.execute("SELECT * FROM categories WHERE id=%s", (cid,)).fetchone()
        if not row:
            flash('Category not found', 'error')
            return redirect(url_for('categories'))
        old_name = row['name']
        db.execute("UPDATE categories SET name=%s, description=%s WHERE id=%s", (new_name, new_desc, cid))
        if row['parent_id']:
            # It's a subcategory — rename it in all products too
            db.execute("UPDATE products SET subcategory=%s WHERE subcategory=%s", (new_name, old_name))
        else:
            # It's a top-level category — rename it in all products and update catalog order
            db.execute("UPDATE products SET category=%s WHERE category=%s", (new_name, old_name))
            db.execute("UPDATE catalog_category_order SET category=%s WHERE category=%s", (new_name, old_name))
        db.commit()
    flash('Renamed successfully', 'success')
    return redirect(url_for('categories'))


@app.route('/categories/delete/<int:cid>', methods=['POST'])
@admin_required
def delete_category(cid):
    with get_db() as db:
        db.execute("DELETE FROM categories WHERE id=%s", (cid,))
        db.commit()
    flash('Category deleted', 'success')
    return redirect(url_for('categories'))


def _save_photo(file):
    if file and file.filename and allowed_file(file.filename):
        from PIL import Image
        import time
        filename = f"{int(time.time())}.jpg"
        path = os.path.join(UPLOAD_FOLDER, filename)
        img = Image.open(file.stream).convert('RGB')
        img.thumbnail((1200, 1200), Image.LANCZOS)
        img.save(path, 'JPEG', quality=82, optimize=True)
        return filename
    return None


def _product_fields(form):
    l = float(form.get('length') or 0)
    w = float(form.get('width') or 0)
    h = float(form.get('height') or 0)
    cbm = round(l * w * h / 1_000_000, 6)  # cm → m³
    return (
        form.get('sku', ''), form.get('barcode', ''), form['name'],
        form.get('category', ''), form.get('subcategory', ''), form.get('unit', 'pcs'),
        float(form.get('cost', 0)), float(form.get('price', 0)),
        float(form.get('min_stock', 0)), form.get('description', ''),
        l, w, h, float(form.get('weight') or 0), cbm,
        float(form.get('carton_qty') or 0),
        float(form.get('ctn_price') or 0),
        float(form.get('china_price') or 0),
        (form.get('china_currency') or 'RMB').upper(),
    )


@app.route('/products/add', methods=['GET', 'POST'])
@admin_required
def add_product():
    with get_db() as db:
        warehouses = db.execute("SELECT * FROM warehouses").fetchall()
        cats = db.execute("SELECT * FROM categories WHERE parent_id IS NULL ORDER BY name").fetchall()
        subcats = db.execute("SELECT * FROM categories WHERE parent_id IS NOT NULL ORDER BY name").fetchall()
    if request.method == 'POST':
        photo = _save_photo(request.files.get('photo'))
        fields = _product_fields(request.form)
        with get_db() as db:
            cur = db.execute(
                "INSERT INTO products(sku,barcode,name,category,subcategory,unit,cost,price,min_stock,description,length,width,height,weight,cbm,carton_qty,ctn_price,china_price,china_currency,photo) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                fields + (photo or '',)
            )
            pid = cur.lastrowid
            for w in warehouses:
                db.execute("INSERT INTO stock VALUES(%s,%s,%s)", (pid, w['id'], float(request.form.get(f'stock_{w["id"]}', 0))))
            db.commit()
        flash('Product added', 'success')
        next_url = request.form.get('next') or url_for('products')
        return redirect(next_url)
    next_url = request.args.get('next') or request.referrer or url_for('products')
    return render_template('product_form.html', product=None, warehouses=warehouses, stock={}, cats=cats, subcats=subcats, next_url=next_url)


@app.route('/products/edit/<int:pid>', methods=['GET', 'POST'])
@admin_required
def edit_product(pid):
    with get_db() as db:
        p = db.execute("SELECT * FROM products WHERE id=%s", (pid,)).fetchone()
        warehouses = db.execute("SELECT * FROM warehouses").fetchall()
        cats = db.execute("SELECT * FROM categories WHERE parent_id IS NULL ORDER BY name").fetchall()
        subcats = db.execute("SELECT * FROM categories WHERE parent_id IS NOT NULL ORDER BY name").fetchall()
        stock = {r['warehouse_id']: r['qty'] for r in db.execute("SELECT * FROM stock WHERE product_id=%s", (pid,)).fetchall()}
        if request.method == 'POST':
            photo = _save_photo(request.files.get('photo'))
            fields = _product_fields(request.form)
            photo_val = photo if photo else (dict(p).get('photo') or '')
            db.execute(
                "UPDATE products SET sku=%s,barcode=%s,name=%s,category=%s,subcategory=%s,unit=%s,cost=%s,price=%s,min_stock=%s,description=%s,length=%s,width=%s,height=%s,weight=%s,cbm=%s,carton_qty=%s,ctn_price=%s,china_price=%s,china_currency=%s,photo=%s WHERE id=%s",
                fields + (photo_val, pid)
            )
            for w in warehouses:
                db.execute("INSERT INTO stock VALUES(%s,%s,%s) ON CONFLICT (product_id, warehouse_id) DO UPDATE SET qty=EXCLUDED.qty", (pid, w['id'], float(request.form.get(f'stock_{w["id"]}', 0))))
            db.commit()
            flash('Product updated', 'success')
            next_url = request.form.get('next') or url_for('products')
            return redirect(next_url)
    next_url = request.args.get('next') or request.referrer or url_for('products')
    return render_template('product_form.html', product=p, warehouses=warehouses, stock=stock, cats=cats, subcats=subcats, next_url=next_url)


@app.route('/products/duplicate/<int:pid>', methods=['POST'])
@admin_required
def duplicate_product(pid):
    with get_db() as db:
        p = db.execute("SELECT * FROM products WHERE id=%s", (pid,)).fetchone()
        cur = db.execute(
            "INSERT INTO products(sku,barcode,name,category,subcategory,unit,cost,price,min_stock,description,length,width,height,weight,cbm,carton_qty,ctn_price,china_price,china_currency,photo) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            ('', p['barcode'], p['name'] + ' (Copy)', p['category'], p['subcategory'] or '', p['unit'], p['cost'], p['price'],
             p['min_stock'], p['description'], p['length'], p['width'], p['height'],
             p['weight'], p['cbm'], p['carton_qty'],
             (dict(p).get('ctn_price') or 0),
             (dict(p).get('china_price') or 0),
             (dict(p).get('china_currency') or 'RMB'),
             p['photo']))
        db.commit()
        new_id = cur.lastrowid
    flash('Product duplicated — update the name and SKU as needed', 'success')
    return redirect(url_for('edit_product', pid=new_id))


@app.route('/products/delete/<int:pid>', methods=['POST'])
@admin_required
def delete_product(pid):
    try:
        with get_db() as db:
            db.execute("DELETE FROM products WHERE id=%s", (pid,))
            db.commit()
        flash('Product deleted', 'success')
    except Exception:
        flash('Cannot delete — this product is used in existing sales, purchases, or other documents.', 'error')
    return redirect(url_for('products'))


@app.route('/products/bulk', methods=['POST'])
@admin_required
def bulk_products():
    ids = request.form.getlist('ids[]')
    action = request.form.get('action')
    if not ids:
        return redirect(url_for('products'))
    with get_db() as db:
        if action == 'delete':
            deleted, failed = 0, 0
            for pid in ids:
                try:
                    db.execute("DELETE FROM products WHERE id=%s", (pid,))
                    deleted += 1
                except Exception:
                    failed += 1
            db.commit()
            flash(f'{deleted} product(s) deleted' + (f', {failed} skipped (in use)' if failed else ''), 'success')
        elif action == 'edit':
            fields, params = [], []
            if request.form.get('category') not in (None, ''):
                fields.append('category=%s'); params.append(request.form['category'])
            if request.form.get('subcategory') is not None:
                fields.append('subcategory=%s'); params.append(request.form.get('subcategory', ''))
            price_adj = request.form.get('price_adj', '').strip()
            if price_adj:
                adj = float(price_adj)
                fields.append('price=ROUND(price*(1+%s/100),3)'); params.append(adj)
                fields.append('cost=ROUND(cost*(1+%s/100),3)'); params.append(adj)
            if request.form.get('min_stock', '').strip():
                fields.append('min_stock=%s'); params.append(float(request.form['min_stock']))
            if fields:
                for pid in ids:
                    db.execute(f"UPDATE products SET {','.join(fields)} WHERE id=%s", params + [pid])
                db.commit()
            flash(f'{len(ids)} product(s) updated', 'success')
    return redirect(url_for('products'))


@app.route('/products/import-template')
@admin_required
def products_import_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    with get_db() as db:
        categories = [r['category'] for r in db.execute("SELECT DISTINCT category FROM products WHERE category!='' ORDER BY category").fetchall()]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'
    headers = ['name*', 'sku', 'barcode', 'category', 'unit', 'cost', 'price', 'min_stock', 'length_cm', 'width_cm', 'height_cm', 'weight_kg', 'carton_qty', 'description']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9E1F2')
    ws.append(['Example Product', 'SKU001', '1234567890123', categories[0] if categories else '', 'pcs', 10.00, 15.00, 5, 30, 20, 10, 0.5, 12, 'Optional description'])
    if categories:
        # Put category list on a hidden sheet and reference it for dropdown
        ref_ws = wb.create_sheet('_categories')
        for i, cat in enumerate(categories, 1):
            ref_ws.cell(row=i, column=1, value=cat)
        ref_ws.sheet_state = 'hidden'
        cat_col = headers.index('category') + 1
        from openpyxl.utils import get_column_letter
        cat_letter = get_column_letter(cat_col)
        dv = DataValidation(
            type='list',
            formula1=f'_categories!$A$1:$A${len(categories)}',
            allow_blank=True,
            showDropDown=False
        )
        dv.sqref = f'{cat_letter}2:{cat_letter}1000'
        ws.add_data_validation(dv)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), 12) + 2
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name='products_import_template.xlsx', as_attachment=True)


@app.route('/products/import', methods=['GET', 'POST'])
@admin_required
def import_products():
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename.endswith('.xlsx'):
            flash('Please upload an .xlsx file', 'error')
            return redirect(url_for('import_products'))
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            flash('File is empty', 'error')
            return redirect(url_for('import_products'))
        # Normalise header names
        raw_headers = [str(h).strip().lower().rstrip('*').replace(' ', '_') if h else '' for h in rows[0]]
        col = {h: i for i, h in enumerate(raw_headers)}
        def get(row, key, default=''):
            i = col.get(key)
            v = row[i] if i is not None and i < len(row) else None
            return v if v is not None else default
        added, skipped = 0, 0
        with get_db() as db:
            for row in rows[1:]:
                name = str(get(row, 'name', '') or '').strip()
                if not name:
                    skipped += 1
                    continue
                l = float(get(row, 'length_cm', 0) or 0)
                w = float(get(row, 'width_cm', 0) or 0)
                h = float(get(row, 'height_cm', 0) or 0)
                cbm = round(l * w * h / 1_000_000, 6)
                db.execute(
                    "INSERT INTO products(name,sku,barcode,category,unit,cost,price,min_stock,length,width,height,weight,cbm,carton_qty,description) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (name, str(get(row, 'sku', '') or ''), str(get(row, 'barcode', '') or ''),
                     str(get(row, 'category', '') or ''), str(get(row, 'unit', 'pcs') or 'pcs'),
                     float(get(row, 'cost', 0) or 0), float(get(row, 'price', 0) or 0),
                     float(get(row, 'min_stock', 0) or 0),
                     l, w, h, float(get(row, 'weight_kg', 0) or 0), cbm,
                     float(get(row, 'carton_qty', 0) or 0),
                     str(get(row, 'description', '') or '')))
                added += 1
            db.commit()
        flash(f'Imported {added} product(s)' + (f', skipped {skipped} empty rows' if skipped else ''), 'success')
        return redirect(url_for('products'))
    return render_template('product_import.html')


@app.route('/products/barcode')
def barcode_lookup():
    bc = request.args.get('q', '')
    with get_db() as db:
        p = db.execute("SELECT * FROM products WHERE barcode=%s OR sku=%s", (bc, bc)).fetchone()
    if p:
        return jsonify({'found': True, 'id': p['id'], 'name': p['name'], 'price': p['price'],
                        'sku': p['sku'], 'unit': p['unit'], 'carton_qty': p['carton_qty'] or 0})
    return jsonify({'found': False})


# ── Write-offs ────────────────────────────────────────────────────────────────

@app.route('/writeoffs')
def writeoffs():
    with get_db() as db:
        rows = db.execute("""SELECT w.*, wh.name as wh_name FROM writeoffs w
            LEFT JOIN warehouses wh ON wh.id=w.warehouse_id ORDER BY doc_date DESC""").fetchall()
    return render_template('writeoffs.html', writeoffs=rows)


@app.route('/writeoffs/add', methods=['GET', 'POST'])
@admin_required
def add_writeoff():
    with get_db() as db:
        warehouses = db.execute("SELECT * FROM warehouses").fetchall()
        products = [dict(r) for r in db.execute("SELECT p.*, COALESCE((SELECT SUM(qty) FROM stock WHERE product_id=p.id), 0) as stock_total FROM products p ORDER BY p.name").fetchall()]
        if request.method == 'POST':
            pids = request.form.getlist('product_id[]')
            qtys = request.form.getlist('qty[]')
            wid = int(request.form['warehouse_id'])
            errors = []
            for pid, qty in zip(pids, qtys):
                if pid and float(qty) > 0:
                    avail = db.execute("SELECT qty FROM stock WHERE product_id=%s AND warehouse_id=%s", (int(pid), wid)).fetchone()
                    if not avail or avail['qty'] < float(qty):
                        pname = db.execute("SELECT name FROM products WHERE id=%s", (int(pid),)).fetchone()['name']
                        errors.append(f'Not enough stock for {pname}')
            if errors:
                for e in errors: flash(e, 'error')
                return render_template('writeoff_form.html', warehouses=warehouses, products=products, today=date.today().isoformat())
            num = next_num('WO', 'writeoffs')
            cur = db.execute("INSERT INTO writeoffs(num,doc_date,warehouse_id,reason,notes,total_cost) VALUES(%s,%s,%s,%s,%s,0)",
                (num, request.form['doc_date'], wid, request.form.get('reason',''), request.form.get('notes','')))
            wo_id = cur.lastrowid
            total_cost = 0
            for pid, qty in zip(pids, qtys):
                if pid and float(qty) > 0:
                    cost = db.execute("SELECT cost FROM products WHERE id=%s", (int(pid),)).fetchone()['cost']
                    db.execute("INSERT INTO writeoff_items(writeoff_id,product_id,qty,cost) VALUES(%s,%s,%s,%s)", (wo_id, int(pid), float(qty), cost))
                    db.execute("UPDATE stock SET qty=qty-%s WHERE product_id=%s AND warehouse_id=%s", (float(qty), int(pid), wid))
                    total_cost += float(qty) * cost
            db.execute("UPDATE writeoffs SET total_cost=%s WHERE id=%s", (total_cost, wo_id))
            db.commit()
            flash(f'Write-off {num} created', 'success')
            return redirect(url_for('writeoffs'))
    return render_template('writeoff_form.html', warehouses=warehouses, products=products, today=date.today().isoformat())


@app.route('/writeoffs/<int:wid>')
def view_writeoff(wid):
    with get_db() as db:
        w = db.execute("SELECT wo.*, wh.name as wh_name FROM writeoffs wo LEFT JOIN warehouses wh ON wh.id=wo.warehouse_id WHERE wo.id=%s", (wid,)).fetchone()
        items = db.execute("SELECT wi.*, p.name as product_name, p.unit FROM writeoff_items wi JOIN products p ON p.id=wi.product_id WHERE wi.writeoff_id=%s", (wid,)).fetchall()
    return render_template('writeoff_view.html', writeoff=w, items=items)


# ── Inventory Counts ──────────────────────────────────────────────────────────

@app.route('/inventory')
def inventory_counts():
    with get_db() as db:
        rows = db.execute("""SELECT ic.*, wh.name as wh_name FROM inventory_counts ic
            LEFT JOIN warehouses wh ON wh.id=ic.warehouse_id ORDER BY doc_date DESC""").fetchall()
    return render_template('inventory_counts.html', counts=rows)


@app.route('/inventory/add', methods=['GET', 'POST'])
@admin_required
def add_inventory_count():
    with get_db() as db:
        warehouses = db.execute("SELECT * FROM warehouses").fetchall()
        if request.method == 'POST':
            wid = int(request.form['warehouse_id'])
            num = next_num('INV', 'inventory_counts')
            cur = db.execute("INSERT INTO inventory_counts(num,doc_date,warehouse_id,status,notes) VALUES(%s,%s,%s,%s,%s)",
                (num, request.form['doc_date'], wid, 'draft', request.form.get('notes','')))
            count_id = cur.lastrowid
            prods = db.execute("SELECT p.id, COALESCE(s.qty,0) as qty FROM products p LEFT JOIN stock s ON s.product_id=p.id AND s.warehouse_id=%s", (wid,)).fetchall()
            for p in prods:
                db.execute("INSERT INTO inventory_count_items(count_id,product_id,qty_system,qty_actual) VALUES(%s,%s,%s,%s)",
                    (count_id, p['id'], p['qty'], p['qty']))
            db.commit()
            return redirect(url_for('edit_inventory_count', cid=count_id))
    return render_template('inventory_count_form.html', warehouses=warehouses, today=date.today().isoformat())


@app.route('/inventory/<int:cid>', methods=['GET', 'POST'])
@admin_required
def edit_inventory_count(cid):
    with get_db() as db:
        count = db.execute("SELECT ic.*, wh.name as wh_name FROM inventory_counts ic LEFT JOIN warehouses wh ON wh.id=ic.warehouse_id WHERE ic.id=%s", (cid,)).fetchone()
        items = db.execute("SELECT ici.*, p.name as product_name, p.sku, p.unit FROM inventory_count_items ici JOIN products p ON p.id=ici.product_id WHERE ici.count_id=%s", (cid,)).fetchall()
        if request.method == 'POST' and count['status'] == 'draft':
            for item in items:
                actual = float(request.form.get(f'actual_{item["id"]}', item['qty_system']))
                db.execute("UPDATE inventory_count_items SET qty_actual=%s WHERE id=%s", (actual, item['id']))
            if request.form.get('action') == 'apply':
                for item in items:
                    actual = float(request.form.get(f'actual_{item["id"]}', item['qty_system']))
                    db.execute("INSERT INTO stock(product_id,warehouse_id,qty) VALUES(%s,%s,%s) ON CONFLICT (product_id, warehouse_id) DO UPDATE SET qty=EXCLUDED.qty",
                        (item['product_id'], count['warehouse_id'], actual))
                db.execute("UPDATE inventory_counts SET status='applied' WHERE id=%s", (cid,))
                db.commit()
                flash('Inventory count applied — stock updated', 'success')
                return redirect(url_for('inventory_counts'))
            db.commit()
            flash('Count saved', 'success')
            return redirect(url_for('edit_inventory_count', cid=cid))
    return render_template('inventory_count_edit.html', count=count, items=items)


@app.route('/inventory/<int:cid>/delete', methods=['POST'])
@admin_required
def delete_inventory_count(cid):
    with get_db() as db:
        db.execute("DELETE FROM inventory_count_items WHERE count_id=%s", (cid,))
        db.execute("DELETE FROM inventory_counts WHERE id=%s", (cid,))
        db.commit()
    flash('Inventory count deleted', 'success')
    return redirect(url_for('inventory_counts'))


# ── Warehouses ───────────────────────────────────────────────────────────────

@app.route('/warehouses')
def warehouses():
    with get_db() as db:
        result = []
        for w in db.execute("SELECT * FROM warehouses").fetchall():
            val = db.execute("SELECT COALESCE(SUM(s.qty*p.cost),0) AS s FROM stock s JOIN products p ON p.id=s.product_id WHERE s.warehouse_id=%s", (w['id'],)).fetchone()['s']
            count = db.execute("SELECT COUNT(*) AS n FROM stock WHERE warehouse_id=%s AND qty>0", (w['id'],)).fetchone()['n']
            result.append({'wh': w, 'value': val, 'product_count': count})
    return render_template('warehouses.html', warehouses=result)


@app.route('/warehouses/add', methods=['GET', 'POST'])
@admin_required
def add_warehouse():
    if request.method == 'POST':
        with get_db() as db:
            cur = db.execute("INSERT INTO warehouses(name,location) VALUES(%s,%s)", (request.form['name'], request.form.get('location','')))
            wid = cur.lastrowid
            for p in db.execute("SELECT id FROM products").fetchall():
                db.execute("INSERT INTO stock VALUES(%s,%s,0) ON CONFLICT DO NOTHING", (p['id'], wid))
            db.commit()
        flash('Warehouse added', 'success')
        return redirect(url_for('warehouses'))
    return render_template('warehouse_form.html', wh=None)


@app.route('/warehouses/edit/<int:wid>', methods=['GET', 'POST'])
@admin_required
def edit_warehouse(wid):
    with get_db() as db:
        wh = db.execute("SELECT * FROM warehouses WHERE id=%s", (wid,)).fetchone()
        if request.method == 'POST':
            db.execute("UPDATE warehouses SET name=%s,location=%s WHERE id=%s", (request.form['name'], request.form.get('location',''), wid))
            db.commit()
            flash('Warehouse updated', 'success')
            return redirect(url_for('warehouses'))
    return render_template('warehouse_form.html', wh=wh)


# ── Customers ─────────────────────────────────────────────────────────────────

@app.route('/customers')
def customers():
    q = request.args.get('q', '')
    ctype = request.args.get('type', '')
    with get_db() as db:
        base = "SELECT * FROM contacts WHERE 1=1"
        params = []
        if q:
            base += " AND (name LIKE %s OR company LIKE %s OR email LIKE %s OR phone LIKE %s OR postcode LIKE %s OR city LIKE %s OR address LIKE %s)"
            like = f'%{q}%'
            params += [like, like, like, like, like, like, like]
        if ctype:
            base += " AND type=%s"
            params.append(ctype)
        contacts = db.execute(base + " ORDER BY name", params).fetchall()
        rows = []
        for c in contacts:
            outstanding = db.execute(
                "SELECT COALESCE(SUM(total),0) AS s FROM sales WHERE (customer_id=%s OR (customer_id IS NULL AND customer=%s)) AND paid=0 AND archived=0 AND status='completed'",
                (c['id'], c['name'])
            ).fetchone()['s']
            rows.append({'contact': dict(c), 'outstanding': outstanding})
    return render_template('customers.html', customers=rows, q=q, ctype=ctype)


@app.route('/customers/<int:cid>')
def view_customer(cid):
    date_from = request.args.get('date_from', '')
    date_till = request.args.get('date_till', '')
    with get_db() as db:
        c = db.execute("SELECT * FROM contacts WHERE id=%s", (cid,)).fetchone()
        history = db.execute("SELECT * FROM contact_history WHERE contact_id=%s ORDER BY event_date DESC", (cid,)).fetchall()
        cname = c['name']
        sales = db.execute(
            "SELECT * FROM sales WHERE (customer_id=%s OR (customer_id IS NULL AND customer=%s)) AND archived=0 ORDER BY doc_date DESC LIMIT 10",
            (cid, cname)).fetchall()
        purchases = db.execute("SELECT * FROM purchases WHERE supplier_id=%s ORDER BY doc_date DESC LIMIT 10", (cid,)).fetchall()
        outstanding = db.execute(
            "SELECT COALESCE(SUM(total),0) AS s FROM sales WHERE (customer_id=%s OR (customer_id IS NULL AND customer=%s)) AND paid=0 AND archived=0 AND status='completed'",
            (cid, cname)
        ).fetchone()['s']

        # Products purchased with optional date filter
        date_params = [cid, cname]
        date_where = ""
        if date_from:
            date_where += " AND s.doc_date >= %s"
            date_params.append(date_from)
        if date_till:
            date_where += " AND s.doc_date <= %s"
            date_params.append(date_till)
        products_bought = db.execute(f"""
            SELECT pr.name as product_name, pr.sku, pr.unit,
                   SUM(si.qty) as total_qty,
                   SUM(si.qty * si.price * (1 - si.discount_pct/100.0)) as total_value
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            JOIN products pr ON pr.id = si.product_id
            WHERE (s.customer_id=%s OR (s.customer_id IS NULL AND s.customer=%s))
              AND s.archived=0 AND s.status='completed'
              {date_where}
            GROUP BY si.product_id, pr.name, pr.sku, pr.unit
            ORDER BY total_qty DESC
        """, date_params).fetchall()

    return render_template('customer_view.html', customer=c, history=history, sales=sales,
                           purchases=purchases, outstanding=outstanding,
                           products_bought=products_bought, date_from=date_from, date_till=date_till)


@app.route('/customers/<int:cid>/note', methods=['POST'])
@admin_required
def add_customer_note(cid):
    with get_db() as db:
        db.execute('INSERT INTO contact_history(contact_id,event_date,event_type,"desc") VALUES(%s,%s,%s,%s)',
            (cid, date.today().isoformat(), request.form.get('type','note'), request.form['desc']))
        db.commit()
    flash('Note added', 'success')
    return redirect(url_for('view_customer', cid=cid))


@app.route('/customers/add', methods=['GET', 'POST'])
@admin_required
def add_customer():
    if request.method == 'POST':
        with get_db() as db:
            db.execute("INSERT INTO contacts(name,type,company,vat_number,phone,email,address,address2,city,postcode,country,balance,discount_pct,currency,notes) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (request.form['name'], request.form['type'], request.form.get('company',''),
                 request.form.get('vat_number',''), request.form.get('phone',''),
                 request.form.get('email',''), request.form.get('address',''),
                 request.form.get('address2',''), request.form.get('city',''),
                 request.form.get('postcode',''), request.form.get('country',''),
                 float(request.form.get('balance',0)), float(request.form.get('discount_pct',0)),
                 request.form.get('currency','GBP'), request.form.get('notes','')))
            db.commit()
        flash('Customer added', 'success')
        return redirect(url_for('customers'))
    return render_template('customer_form.html', customer=None)


@app.route('/customers/edit/<int:cid>', methods=['GET', 'POST'])
@admin_required
def edit_customer(cid):
    with get_db() as db:
        c = db.execute("SELECT * FROM contacts WHERE id=%s", (cid,)).fetchone()
        if request.method == 'POST':
            new_name = request.form['name']
            new_company = request.form.get('company', '')
            old_name = c['name']
            old_company = c['company'] or ''
            old_label = old_company if old_company else old_name
            new_label = new_company if new_company else new_name
            db.execute("UPDATE contacts SET name=%s,type=%s,company=%s,vat_number=%s,phone=%s,email=%s,address=%s,address2=%s,city=%s,postcode=%s,country=%s,balance=%s,discount_pct=%s,currency=%s,notes=%s WHERE id=%s",
                (new_name, request.form['type'], new_company,
                 request.form.get('vat_number',''), request.form.get('phone',''),
                 request.form.get('email',''), request.form.get('address',''),
                 request.form.get('address2',''), request.form.get('city',''),
                 request.form.get('postcode',''), request.form.get('country',''),
                 float(request.form.get('balance',0)), float(request.form.get('discount_pct',0)),
                 request.form.get('currency','GBP'), request.form.get('notes',''), cid))
            if new_label != old_label:
                db.execute("UPDATE sales SET customer=%s WHERE customer_id=%s OR (customer_id IS NULL AND customer=%s)", (new_label, cid, old_label))
                db.execute("UPDATE quotes SET customer=%s WHERE customer_id=%s OR (customer_id IS NULL AND customer=%s)", (new_label, cid, old_label))
            db.commit()
            flash('Customer updated', 'success')
            return redirect(url_for('customers'))
    return render_template('customer_form.html', customer=c)


@app.route('/customers/delete/<int:cid>', methods=['POST'])
@admin_required
def delete_customer(cid):
    with get_db() as db:
        db.execute("DELETE FROM contacts WHERE id=%s", (cid,))
        db.commit()
    flash('Customer deleted', 'success')
    return redirect(url_for('customers'))


# ── Purchases ────────────────────────────────────────────────────────────────

@app.route('/purchases')
def purchases():
    with get_db() as db:
        rows = db.execute("SELECT p.*, w.name as wh_name FROM purchases p LEFT JOIN warehouses w ON w.id=p.warehouse_id ORDER BY doc_date DESC").fetchall()
    return render_template('purchases.html', purchases=rows)


@app.route('/purchases/bulk', methods=['POST'])
@admin_required
def bulk_purchases():
    ids = request.form.getlist('ids[]')
    action = request.form.get('action')
    if not ids:
        return redirect(url_for('purchases'))
    count = 0
    with get_db() as db:
        for pid in ids:
            if action == 'delete':
                db.execute("DELETE FROM purchases WHERE id=%s", (pid,))
                count += 1
        db.commit()
    flash(f'{count} purchase(s) deleted', 'success')
    return redirect(url_for('purchases'))


@app.route('/purchases/add', methods=['GET', 'POST'])
@admin_required
def add_purchase():
    with get_db() as db:
        suppliers  = [dict(r) for r in db.execute("SELECT * FROM contacts WHERE type='supplier' ORDER BY name").fetchall()]
        warehouses = [dict(r) for r in db.execute("SELECT * FROM warehouses").fetchall()]
        products   = [dict(r) for r in db.execute("SELECT p.*, COALESCE((SELECT SUM(qty) FROM stock WHERE product_id=p.id), 0) as stock_total FROM products p ORDER BY p.name").fetchall()]
        if request.method == 'POST':
            pids    = request.form.getlist('product_id[]')
            qtys    = request.form.getlist('qty[]')
            prices  = request.form.getlist('price[]')
            tax_pct = float(request.form.get('tax_pct', 0))
            discount= float(request.form.get('global_discount', 0))
            line_total = sum(float(q)*float(p) for q,p in zip(qtys,prices) if q and p)
            disc_amt   = line_total * discount / 100
            subtotal   = line_total - disc_amt
            tax_amount = subtotal * tax_pct / 100
            total      = subtotal + tax_amount
            num         = next_num('PUR', 'purchases')
            wid         = int(request.form['warehouse_id'])
            supplier_id = request.form.get('supplier_id') or None
            cur = db.execute(
                "INSERT INTO purchases(num,doc_date,supplier,supplier_id,warehouse_id,subtotal,total,tax_pct,tax_amount,discount,currency,status,notes) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (num, request.form['doc_date'], request.form['supplier'], supplier_id, wid,
                 subtotal, total, tax_pct, tax_amount, discount,
                 request.form.get('currency','GBP'), 'received', request.form.get('notes','')))
            purchase_id = cur.lastrowid
            for pid, qty, price in zip(pids, qtys, prices):
                if pid and float(qty or 0) > 0:
                    db.execute("INSERT INTO purchase_items(purchase_id,product_id,qty,price) VALUES(%s,%s,%s,%s)",
                               (purchase_id, int(pid), float(qty), float(price)))
                    db.execute("INSERT INTO stock(product_id,warehouse_id,qty) VALUES(%s,%s,%s) ON CONFLICT(product_id,warehouse_id) DO UPDATE SET qty=qty+excluded.qty",
                               (int(pid), wid, float(qty)))
            if supplier_id:
                db.execute('INSERT INTO contact_history(contact_id,event_date,event_type,"desc") VALUES(%s,%s,\'purchase\',%s)',
                           (supplier_id, request.form['doc_date'], f'Purchase {num} — £{total:,.2f}'))
            db.commit()
            flash(f'Purchase {num} created', 'success')
            return redirect(url_for('purchases'))
    return render_template('purchase_form.html', purchase=None, suppliers=suppliers,
                           warehouses=warehouses, products=products, items=[],
                           today=date.today().isoformat())


@app.route('/purchases/<int:pid>/edit', methods=['GET', 'POST'])
@admin_required
def edit_purchase(pid):
    with get_db() as db:
        purchase   = dict(db.execute("SELECT * FROM purchases WHERE id=%s", (pid,)).fetchone())
        suppliers  = [dict(r) for r in db.execute("SELECT * FROM contacts WHERE type='supplier' ORDER BY name").fetchall()]
        warehouses = [dict(r) for r in db.execute("SELECT * FROM warehouses").fetchall()]
        products   = [dict(r) for r in db.execute("SELECT p.*, COALESCE((SELECT SUM(qty) FROM stock WHERE product_id=p.id), 0) as stock_total FROM products p ORDER BY p.name").fetchall()]
        items      = [dict(r) for r in db.execute("SELECT * FROM purchase_items WHERE purchase_id=%s", (pid,)).fetchall()]
        if request.method == 'POST':
            pids     = request.form.getlist('product_id[]')
            qtys     = request.form.getlist('qty[]')
            prices   = request.form.getlist('price[]')
            tax_pct  = float(request.form.get('tax_pct', 0))
            discount = float(request.form.get('global_discount', 0))
            line_total = sum(float(q)*float(p) for q,p in zip(qtys,prices) if q and p)
            disc_amt   = line_total * discount / 100
            subtotal   = line_total - disc_amt
            tax_amount = subtotal * tax_pct / 100
            total      = subtotal + tax_amount
            wid        = int(request.form['warehouse_id'])
            supplier_id= request.form.get('supplier_id') or None
            # Reverse old stock
            old_items = db.execute("SELECT * FROM purchase_items WHERE purchase_id=%s", (pid,)).fetchall()
            old_wid   = purchase['warehouse_id']
            for oi in old_items:
                db.execute("UPDATE stock SET qty=qty-%s WHERE product_id=%s AND warehouse_id=%s",
                           (oi['qty'], oi['product_id'], old_wid))
            db.execute("DELETE FROM purchase_items WHERE purchase_id=%s", (pid,))
            db.execute("""UPDATE purchases SET doc_date=%s,supplier=%s,supplier_id=%s,warehouse_id=%s,
                          subtotal=%s,total=%s,tax_pct=%s,tax_amount=%s,discount=%s,currency=%s,notes=%s
                          WHERE id=%s""",
                       (request.form['doc_date'], request.form['supplier'], supplier_id, wid,
                        subtotal, total, tax_pct, tax_amount, discount,
                        request.form.get('currency','GBP'), request.form.get('notes',''), pid))
            for p_id, qty, price in zip(pids, qtys, prices):
                if p_id and float(qty or 0) > 0:
                    db.execute("INSERT INTO purchase_items(purchase_id,product_id,qty,price) VALUES(%s,%s,%s,%s)",
                               (pid, int(p_id), float(qty), float(price)))
                    db.execute("INSERT INTO stock(product_id,warehouse_id,qty) VALUES(%s,%s,%s) ON CONFLICT(product_id,warehouse_id) DO UPDATE SET qty=qty+excluded.qty",
                               (int(p_id), wid, float(qty)))
            db.commit()
            flash(f'Purchase {purchase["num"]} updated', 'success')
            return redirect(url_for('view_purchase', pid=pid))
    return render_template('purchase_form.html', purchase=purchase, suppliers=suppliers,
                           warehouses=warehouses, products=products, items=items,
                           today=date.today().isoformat())


@app.route('/purchases/<int:pid>')
def view_purchase(pid):
    with get_db() as db:
        p = db.execute("SELECT pu.*, w.name as wh_name FROM purchases pu LEFT JOIN warehouses w ON w.id=pu.warehouse_id WHERE pu.id=%s", (pid,)).fetchone()
        items = db.execute("SELECT pi.*, pr.name as product_name, pr.unit FROM purchase_items pi JOIN products pr ON pr.id=pi.product_id WHERE pi.purchase_id=%s", (pid,)).fetchall()
        debit_notes_list = db.execute("SELECT * FROM debit_notes WHERE purchase_id=%s ORDER BY doc_date DESC", (pid,)).fetchall()
    return render_template('purchase_view.html', purchase=p, items=items, debit_notes_list=debit_notes_list)


@app.route('/purchases/<int:pid>/export/pdf')
def export_purchase_pdf(pid):
    with get_db() as db:
        p = dict(db.execute("SELECT pu.*, w.name as wh_name FROM purchases pu LEFT JOIN warehouses w ON w.id=pu.warehouse_id WHERE pu.id=%s", (pid,)).fetchone())
        items = [dict(r) for r in db.execute("SELECT pi.*, pr.name as product_name, pr.unit FROM purchase_items pi JOIN products pr ON pr.id=pi.product_id WHERE pi.purchase_id=%s", (pid,)).fetchall()]
    return send_file(io.BytesIO(generate_pdf(f"Purchase {p['num']}", p, items)), mimetype='application/pdf', download_name=f"{p['num']}.pdf", as_attachment=True)


@app.route('/purchases/<int:pid>/export/excel')
def export_purchase_excel(pid):
    with get_db() as db:
        p = dict(db.execute("SELECT pu.*, w.name as wh_name FROM purchases pu LEFT JOIN warehouses w ON w.id=pu.warehouse_id WHERE pu.id=%s", (pid,)).fetchone())
        items = [dict(r) for r in db.execute("SELECT pi.*, pr.name as product_name, pr.unit, pr.carton_qty FROM purchase_items pi JOIN products pr ON pr.id=pi.product_id WHERE pi.purchase_id=%s", (pid,)).fetchall()]
        supplier = None
        if p.get('supplier_id'):
            row = db.execute("SELECT * FROM contacts WHERE id=%s", (p['supplier_id'],)).fetchone()
            if row: supplier = dict(row)
    company = get_settings()
    return send_file(io.BytesIO(generate_excel(p['num'], p, items, doc_type='purchase', contact=supplier, company=company)), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', download_name=f"{p['num']}.xlsx", as_attachment=True)


# ── Supplier Orders ───────────────────────────────────────────────────────────

@app.route('/supplier-orders')
def supplier_orders():
    with get_db() as db:
        rows = db.execute("""
            SELECT o.*, (SELECT COUNT(*) FROM supplier_order_items WHERE order_id=o.id) AS item_count,
                        (SELECT COALESCE(SUM(cbm*ctns),0) FROM supplier_order_items WHERE order_id=o.id) AS total_cbm,
                        (SELECT COALESCE(SUM(ctns),0) FROM supplier_order_items WHERE order_id=o.id) AS total_ctns
            FROM supplier_orders o ORDER BY doc_date DESC, id DESC
        """).fetchall()
    return render_template('supplier_orders.html', orders=rows)


@app.route('/supplier-orders/add', methods=['GET', 'POST'])
@admin_required
def add_supplier_order():
    with get_db() as db:
        suppliers = [dict(r) for r in db.execute("SELECT * FROM contacts WHERE type='supplier' ORDER BY name").fetchall()]
        products = [dict(r) for r in db.execute("SELECT id, name, sku, photo, cbm, carton_qty FROM products ORDER BY name").fetchall()]
        if request.method == 'POST':
            num = next_num('SO', 'supplier_orders')
            supplier_id = request.form.get('supplier_id') or None
            cur = db.execute(
                "INSERT INTO supplier_orders(num,doc_date,supplier,supplier_id,notes,status) VALUES(%s,%s,%s,%s,%s,%s)",
                (num, request.form['doc_date'], request.form.get('supplier',''),
                 supplier_id, request.form.get('notes',''), 'draft'))
            order_id = cur.lastrowid
            _save_supplier_order_items(db, order_id)
            db.commit()
            flash(f'Order {num} created', 'success')
            return redirect(url_for('view_supplier_order', oid=order_id))
    return render_template('supplier_order_form.html', order=None, items=[],
                           suppliers=suppliers, products=products, today=date.today().isoformat())


@app.route('/supplier-orders/<int:oid>/edit', methods=['GET', 'POST'])
@admin_required
def edit_supplier_order(oid):
    with get_db() as db:
        order = db.execute("SELECT * FROM supplier_orders WHERE id=%s", (oid,)).fetchone()
        if not order:
            flash('Order not found', 'error')
            return redirect(url_for('supplier_orders'))
        order = dict(order)
        suppliers = [dict(r) for r in db.execute("SELECT * FROM contacts WHERE type='supplier' ORDER BY name").fetchall()]
        products = [dict(r) for r in db.execute("SELECT id, name, sku, photo, cbm, carton_qty FROM products ORDER BY name").fetchall()]
        items = [dict(r) for r in db.execute("SELECT * FROM supplier_order_items WHERE order_id=%s ORDER BY id", (oid,)).fetchall()]
        if request.method == 'POST':
            supplier_id = request.form.get('supplier_id') or None
            db.execute("""UPDATE supplier_orders SET doc_date=%s,supplier=%s,supplier_id=%s,notes=%s WHERE id=%s""",
                       (request.form['doc_date'], request.form.get('supplier',''),
                        supplier_id, request.form.get('notes',''), oid))
            _save_supplier_order_items(db, oid, replace=True)
            db.commit()
            flash(f'Order {order["num"]} updated', 'success')
            return redirect(url_for('view_supplier_order', oid=oid))
    return render_template('supplier_order_form.html', order=order, items=items,
                           suppliers=suppliers, products=products, today=date.today().isoformat())


def _save_supplier_order_items(db, order_id, replace=False):
    old_photos = set()
    if replace:
        old_photos = set(r['photo'] for r in db.execute(
            "SELECT photo FROM supplier_order_items WHERE order_id=%s", (order_id,)).fetchall() if r['photo'])
        db.execute("DELETE FROM supplier_order_items WHERE order_id=%s", (order_id,))

    names    = request.form.getlist('item_name[]')
    articles = request.form.getlist('item_article[]')
    cbms     = request.form.getlist('item_cbm[]')
    ctnss    = request.form.getlist('item_ctns[]')
    ppcs     = request.form.getlist('item_pcs_per_ctn[]')
    existing = request.form.getlist('existing_photo[]')
    files    = request.files.getlist('item_photo[]')

    used_photos = set()
    for i, name in enumerate(names):
        if not (name or '').strip() and not (articles[i] if i < len(articles) else '').strip():
            continue
        photo = ''
        if i < len(files):
            saved = _save_photo(files[i])
            if saved:
                photo = saved
        if not photo and i < len(existing):
            photo = existing[i]
        if photo:
            used_photos.add(photo)
        db.execute(
            "INSERT INTO supplier_order_items(order_id,name,article,photo,cbm,ctns,pcs_per_ctn) VALUES(%s,%s,%s,%s,%s,%s,%s)",
            (order_id, name, articles[i] if i < len(articles) else '', photo,
             float(cbms[i] or 0) if i < len(cbms) else 0,
             float(ctnss[i] or 0) if i < len(ctnss) else 0,
             float(ppcs[i] or 0) if i < len(ppcs) else 0))

    if replace:
        for stale in old_photos - used_photos:
            try:
                os.remove(os.path.join(UPLOAD_FOLDER, stale))
            except OSError:
                pass


@app.route('/supplier-orders/<int:oid>')
def view_supplier_order(oid):
    with get_db() as db:
        order = db.execute("SELECT * FROM supplier_orders WHERE id=%s", (oid,)).fetchone()
        if not order:
            flash('Order not found', 'error')
            return redirect(url_for('supplier_orders'))
        items = [dict(r) for r in db.execute("SELECT * FROM supplier_order_items WHERE order_id=%s ORDER BY id", (oid,)).fetchall()]
    return render_template('supplier_order_view.html', order=dict(order), items=items)


@app.route('/supplier-orders/<int:oid>/export/excel')
def export_supplier_order_excel(oid):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    with get_db() as db:
        order = db.execute("SELECT * FROM supplier_orders WHERE id=%s", (oid,)).fetchone()
        if not order:
            flash('Order not found', 'error')
            return redirect(url_for('supplier_orders'))
        order = dict(order)
        items = [dict(r) for r in db.execute("SELECT * FROM supplier_order_items WHERE order_id=%s ORDER BY id", (oid,)).fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = order['num'][:31]

    thin = Side(style='thin', color='D0D7DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='1F2937')
    title_font = Font(bold=True, size=16, color='0F172A')
    meta_font = Font(size=11, color='475569')
    head_font = Font(bold=True, size=11, color='FFFFFF')

    ws['A1'] = f"Supplier Order  {order['num']}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')
    ws['A2'] = f"Date: {order['doc_date']}   |   Supplier: {order['supplier'] or '—'}"
    ws['A2'].font = meta_font
    ws.merge_cells('A2:H2')
    if order.get('notes'):
        ws['A3'] = f"Notes: {order['notes']}"
        ws['A3'].font = meta_font
        ws.merge_cells('A3:H3')
        header_row = 5
    else:
        header_row = 4

    headers = ['#', 'Photo', 'Name', 'Article', 'CBM', 'CTNS', 'Pcs/CTN', 'Total CBM']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = head_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.row_dimensions[header_row].height = 22

    widths = [5, 14, 32, 14, 10, 10, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total_ctns = 0.0
    total_cbm = 0.0
    row = header_row + 1
    for idx, it in enumerate(items, start=1):
        ws.row_dimensions[row].height = 60
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal='center', vertical='center')

        photo_path = os.path.join(UPLOAD_FOLDER, it['photo']) if it.get('photo') else None
        if photo_path and os.path.exists(photo_path):
            try:
                with PILImage.open(photo_path) as pim:
                    pim.thumbnail((96, 96), PILImage.LANCZOS)
                    tmp_buf = io.BytesIO()
                    pim.convert('RGB').save(tmp_buf, format='PNG')
                    tmp_buf.seek(0)
                xlimg = XLImage(tmp_buf)
                xlimg.width = 72
                xlimg.height = 72
                ws.add_image(xlimg, f'B{row}')
            except Exception:
                pass

        ctns = float(it.get('ctns') or 0)
        cbm  = float(it.get('cbm') or 0)
        ppc  = float(it.get('pcs_per_ctn') or 0)
        line_cbm = ctns * cbm
        total_ctns += ctns
        total_cbm  += line_cbm

        values = [it.get('name', ''), it.get('article', ''), cbm, ctns, ppc, line_cbm]
        for ci, v in enumerate(values, start=3):
            c = ws.cell(row=row, column=ci, value=v)
            c.alignment = Alignment(vertical='center', horizontal='left' if ci in (3, 4) else 'right', wrap_text=True)
            c.border = border
            if ci == 5 or ci == 8:
                c.number_format = '0.000'
            elif ci in (6, 7):
                c.number_format = '0'
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).border = border
        row += 1

    # Totals row
    tot_font = Font(bold=True, size=11)
    tot_fill = PatternFill('solid', fgColor='F1F5F9')
    ws.cell(row=row, column=1, value='Totals').font = tot_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    for col in range(1, 9):
        cc = ws.cell(row=row, column=col)
        cc.fill = tot_fill
        cc.border = border
        cc.font = tot_font
    ws.cell(row=row, column=6, value=total_ctns).number_format = '0'
    ws.cell(row=row, column=8, value=total_cbm).number_format = '0.000'
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
    for col in (6, 8):
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='right', vertical='center')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=f"{order['num']}.xlsx", as_attachment=True)


@app.route('/supplier-orders/<int:oid>/delete', methods=['POST'])
@admin_required
def delete_supplier_order(oid):
    with get_db() as db:
        photos = db.execute("SELECT photo FROM supplier_order_items WHERE order_id=%s", (oid,)).fetchall()
        for p in photos:
            if p['photo']:
                try:
                    os.remove(os.path.join(UPLOAD_FOLDER, p['photo']))
                except OSError:
                    pass
        db.execute("DELETE FROM supplier_orders WHERE id=%s", (oid,))
        db.commit()
    flash('Order deleted', 'success')
    return redirect(url_for('supplier_orders'))


# ── Sales ─────────────────────────────────────────────────────────────────────

@app.route('/sales')
def sales():
    date_from = request.args.get('date_from', '')
    date_till = request.args.get('date_till', '')
    customer_id = request.args.get('customer_id', '')
    status_filter = request.args.get('status', '')
    with get_db() as db:
        params = []
        where = "WHERE s.archived=0"
        if date_from:
            where += " AND s.doc_date >= %s"
            params.append(date_from)
        if date_till:
            where += " AND s.doc_date <= %s"
            params.append(date_till)
        if customer_id:
            where += " AND s.customer_id = %s"
            params.append(customer_id)
        if status_filter:
            where += " AND s.status = %s"
            params.append(status_filter)
        rows = [dict(r) for r in db.execute(f"""
            SELECT s.*, COALESCE(NULLIF(c.company,''), NULLIF(s.customer,''), c.name) as customer
            FROM sales s LEFT JOIN contacts c ON c.id=s.customer_id
            {where} ORDER BY s.doc_date DESC
        """, params).fetchall()]
        customers = db.execute(
            "SELECT id, COALESCE(NULLIF(company,''), name) as label FROM contacts WHERE type='customer' ORDER BY label"
        ).fetchall()
    return render_template('sales.html', sales=rows, date_from=date_from, date_till=date_till,
                           customer_id=customer_id, customers=customers, status_filter=status_filter)


@app.route('/sales/archive')
def sales_archive():
    with get_db() as db:
        rows = [dict(r) for r in db.execute("""
            SELECT s.*, COALESCE(NULLIF(c.company,''), NULLIF(s.customer,''), c.name) as customer
            FROM sales s LEFT JOIN contacts c ON c.id=s.customer_id
            WHERE s.archived=1 ORDER BY s.doc_date DESC
        """).fetchall()]
    return render_template('sales_archive.html', sales=rows)


@app.route('/sales/add', methods=['GET', 'POST'])
@admin_required
def add_sale():
    with get_db() as db:
        customers = [dict(r) for r in db.execute("SELECT * FROM contacts WHERE type='customer' ORDER BY name").fetchall()]
        warehouses = [dict(r) for r in db.execute("SELECT * FROM warehouses").fetchall()]
        products = [dict(r) for r in db.execute("SELECT p.*, COALESCE((SELECT SUM(qty) FROM stock WHERE product_id=p.id), 0) as stock_total FROM products p ORDER BY p.name").fetchall()]
        discounts = db.execute("SELECT * FROM discounts WHERE active=1").fetchall()
        if request.method == 'POST':
            pids = request.form.getlist('product_id[]')
            wids = request.form.getlist('warehouse_id[]')
            qtys = request.form.getlist('qty[]')
            prices = request.form.getlist('price[]')
            disc_pcts = request.form.getlist('discount_pct[]')
            global_disc = float(request.form.get('global_discount', 0))
            tax_pct = float(request.form.get('tax_pct', 0))
            subtotal = sum(float(q)*float(p)*(1-float(d)/100) for q,p,d in zip(qtys,prices,disc_pcts) if q and p) * (1 - global_disc/100)
            tax_amount = subtotal * tax_pct / 100
            total = subtotal + tax_amount
            num = next_num('SAL', 'sales')
            customer_id = request.form.get('customer_id') or None
            cur = db.execute("INSERT INTO sales(num,doc_date,customer,customer_id,subtotal,total,discount,tax_pct,tax_amount,currency,status,notes) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (num, request.form['doc_date'], request.form['customer'], customer_id, subtotal, total, global_disc, tax_pct, tax_amount,
                 request.form.get('currency','GBP'), 'completed', request.form.get('notes','')))
            sale_id = cur.lastrowid
            for pid, wid, qty, price, disc in zip(pids, wids, qtys, prices, disc_pcts):
                if pid and float(qty) > 0:
                    db.execute("INSERT INTO sale_items(sale_id,product_id,warehouse_id,qty,price,discount_pct) VALUES(%s,%s,%s,%s,%s,%s)",
                        (sale_id, int(pid), int(wid), float(qty), float(price), float(disc)))
                    db.execute("UPDATE stock SET qty=qty-%s WHERE product_id=%s AND warehouse_id=%s", (float(qty), int(pid), int(wid)))
            if customer_id:
                db.execute('INSERT INTO contact_history(contact_id,event_date,event_type,"desc") VALUES(%s,%s,\'sale\',%s)',
                    (customer_id, request.form['doc_date'], f'Sale {num} — £{total:,.2f}'))
            db.commit()
            flash(f'Sale {num} created', 'success')
            return redirect(url_for('sales'))
    return render_template('sale_form.html', customers=customers, warehouses=warehouses, products=products, discounts=discounts, today=date.today().isoformat())


@app.route('/sales/<int:sid>')
def view_sale(sid):
    with get_db() as db:
        s = dict(db.execute("""
            SELECT s.*, COALESCE(NULLIF(c.company,''), NULLIF(s.customer,''), c.name) as customer
            FROM sales s LEFT JOIN contacts c ON c.id=s.customer_id WHERE s.id=%s
        """, (sid,)).fetchone())
        items = db.execute("SELECT si.*, pr.name as product_name, pr.unit, pr.carton_qty, w.name as wh_name FROM sale_items si JOIN products pr ON pr.id=si.product_id JOIN warehouses w ON w.id=si.warehouse_id WHERE si.sale_id=%s", (sid,)).fetchall()
        customer = None
        if s.get('customer_id'):
            row = db.execute("SELECT * FROM contacts WHERE id=%s", (s['customer_id'],)).fetchone()
            if row:
                customer = dict(row)
        payments = db.execute("SELECT * FROM invoice_payments WHERE sale_id=%s ORDER BY payment_date", (sid,)).fetchall()
        paid_total = sum(p['amount'] for p in payments)
        outstanding = max(0, s['total'] - paid_total)
        credit_notes_list = db.execute("SELECT * FROM credit_notes WHERE sale_id=%s ORDER BY doc_date DESC", (sid,)).fetchall()
    company = get_settings()
    return render_template('sale_view.html', sale=s, items=items, customer=customer, company=company,
                           payments=payments, paid_total=paid_total, outstanding=outstanding,
                           credit_notes_list=credit_notes_list)


@app.route('/sales/<int:sid>/edit', methods=['GET', 'POST'])
@admin_required
def edit_sale(sid):
    with get_db() as db:
        sale = db.execute("SELECT * FROM sales WHERE id=%s", (sid,)).fetchone()
        if not sale or sale['status'] == 'cancelled':
            flash('Sale cannot be edited', 'error')
            return redirect(url_for('sales'))
        customers = [dict(r) for r in db.execute("SELECT * FROM contacts WHERE type='customer' ORDER BY name").fetchall()]
        warehouses = [dict(r) for r in db.execute("SELECT * FROM warehouses").fetchall()]
        products = [dict(r) for r in db.execute("SELECT p.*, COALESCE((SELECT SUM(qty) FROM stock WHERE product_id=p.id), 0) as stock_total FROM products p ORDER BY p.name").fetchall()]
        discounts = db.execute("SELECT * FROM discounts WHERE active=1").fetchall()
        items = [dict(r) for r in db.execute("SELECT * FROM sale_items WHERE sale_id=%s", (sid,)).fetchall()]
        if request.method == 'POST':
            pids = request.form.getlist('product_id[]')
            wids = request.form.getlist('warehouse_id[]')
            qtys = request.form.getlist('qty[]')
            prices = request.form.getlist('price[]')
            disc_pcts = request.form.getlist('discount_pct[]')
            # Restore stock from old items before validation
            for item in items:
                db.execute("UPDATE stock SET qty=qty+%s WHERE product_id=%s AND warehouse_id=%s",
                    (item['qty'], item['product_id'], item['warehouse_id']))
            global_disc = float(request.form.get('global_discount', 0))
            tax_pct = float(request.form.get('tax_pct', 0))
            subtotal = sum(float(q)*float(p)*(1-float(d)/100) for q,p,d in zip(qtys,prices,disc_pcts) if q and p) * (1 - global_disc/100)
            tax_amount = subtotal * tax_pct / 100
            total = subtotal + tax_amount
            customer_id = request.form.get('customer_id') or None
            db.execute("UPDATE sales SET doc_date=%s,customer=%s,customer_id=%s,subtotal=%s,total=%s,discount=%s,tax_pct=%s,tax_amount=%s,currency=%s,notes=%s WHERE id=%s",
                (request.form['doc_date'], request.form['customer'], customer_id, subtotal, total, global_disc, tax_pct, tax_amount,
                 request.form.get('currency', 'GBP'), request.form.get('notes', ''), sid))
            db.execute("DELETE FROM sale_items WHERE sale_id=%s", (sid,))
            for pid, wid, qty, price, disc in zip(pids, wids, qtys, prices, disc_pcts):
                if pid and float(qty) > 0:
                    db.execute("INSERT INTO sale_items(sale_id,product_id,warehouse_id,qty,price,discount_pct) VALUES(%s,%s,%s,%s,%s,%s)",
                        (sid, int(pid), int(wid), float(qty), float(price), float(disc)))
                    db.execute("UPDATE stock SET qty=qty-%s WHERE product_id=%s AND warehouse_id=%s", (float(qty), int(pid), int(wid)))
            db.commit()
            flash(f'Sale {sale["num"]} updated', 'success')
            return redirect(url_for('view_sale', sid=sid))
    return render_template('sale_form.html', sale=dict(sale), items=items,
        customers=customers, warehouses=warehouses, products=products, discounts=discounts)


@app.route('/sales/cancel/<int:sid>', methods=['POST'])
@admin_required
def cancel_sale(sid):
    with get_db() as db:
        s = db.execute("SELECT * FROM sales WHERE id=%s", (sid,)).fetchone()
        if s['status'] == 'completed':
            for item in db.execute("SELECT * FROM sale_items WHERE sale_id=%s", (sid,)).fetchall():
                db.execute("UPDATE stock SET qty=qty+%s WHERE product_id=%s AND warehouse_id=%s", (item['qty'], item['product_id'], item['warehouse_id']))
        db.execute("UPDATE sales SET status='cancelled' WHERE id=%s", (sid,))
        db.commit()
    flash('Sale cancelled and stock restored', 'success')
    return redirect(url_for('sales'))


@app.route('/sales/<int:sid>/toggle_paid', methods=['POST'])
@admin_required
def toggle_paid(sid):
    with get_db() as db:
        s = db.execute("SELECT * FROM sales WHERE id=%s", (sid,)).fetchone()
        now_paid = 0 if s['paid'] else 1
        db.execute("UPDATE sales SET paid=%s WHERE id=%s", (now_paid, sid))
        if s['customer_id']:
            # paid → balance increases (debt cleared); unpaid → balance decreases (owes)
            delta = s['total'] if now_paid else -s['total']
            db.execute("UPDATE contacts SET balance=balance+%s WHERE id=%s", (delta, s['customer_id']))
        db.commit()
    return redirect(request.referrer or url_for('sales'))


@app.route('/sales/<int:sid>/archive', methods=['POST'])
@admin_required
def archive_sale(sid):
    with get_db() as db:
        s = db.execute("SELECT * FROM sales WHERE id=%s", (sid,)).fetchone()
        if s['status'] == 'completed':
            for item in db.execute("SELECT * FROM sale_items WHERE sale_id=%s", (sid,)).fetchall():
                db.execute("UPDATE stock SET qty=qty+%s WHERE product_id=%s AND warehouse_id=%s",
                           (item['qty'], item['product_id'], item['warehouse_id']))
        # snapshot product name then detach product_id so products can be deleted
        db.execute("""UPDATE sale_items SET product_name_snap=COALESCE(
            (SELECT name FROM products WHERE products.id=sale_items.product_id), product_name_snap),
            product_id=NULL WHERE sale_id=%s""", (sid,))
        db.execute("UPDATE sales SET archived=1, status='archived' WHERE id=%s", (sid,))
        db.commit()
    flash('Sale archived', 'success')
    return redirect(url_for('sales'))


@app.route('/sales/<int:sid>/delete', methods=['POST'])
@admin_required
def delete_sale(sid):
    with get_db() as db:
        db.execute("DELETE FROM sales WHERE id=%s", (sid,))
        db.commit()
    flash('Sale permanently deleted', 'success')
    return redirect(url_for('sales_archive'))


@app.route('/sales/bulk', methods=['POST'])
@admin_required
def bulk_sales():
    ids = request.form.getlist('ids[]')
    action = request.form.get('action')
    if not ids:
        return redirect(url_for('sales'))
    count = 0
    with get_db() as db:
        for sid in ids:
            s = db.execute("SELECT * FROM sales WHERE id=%s", (sid,)).fetchone()
            if not s:
                continue
            if action == 'archive':
                if s['status'] == 'completed':
                    for item in db.execute("SELECT * FROM sale_items WHERE sale_id=%s", (sid,)).fetchall():
                        db.execute("UPDATE stock SET qty=qty+%s WHERE product_id=%s AND warehouse_id=%s",
                                   (item['qty'], item['product_id'], item['warehouse_id']))
                db.execute("""UPDATE sale_items SET product_name_snap=COALESCE(
                    (SELECT name FROM products WHERE products.id=sale_items.product_id), product_name_snap),
                    product_id=NULL WHERE sale_id=%s""", (sid,))
                db.execute("UPDATE sales SET archived=1, status='archived' WHERE id=%s", (sid,))
                count += 1
            elif action == 'cancel':
                if s['status'] == 'completed':
                    for item in db.execute("SELECT * FROM sale_items WHERE sale_id=%s", (sid,)).fetchall():
                        db.execute("UPDATE stock SET qty=qty+%s WHERE product_id=%s AND warehouse_id=%s",
                                   (item['qty'], item['product_id'], item['warehouse_id']))
                    db.execute("UPDATE sales SET status='cancelled' WHERE id=%s", (sid,))
                    count += 1
            elif action == 'delete':
                db.execute("DELETE FROM sales WHERE id=%s", (sid,))
                count += 1
        db.commit()
    flash(f'{count} sale(s) {action}d', 'success')
    return redirect(url_for('sales_archive') if action in ('archive', 'delete') else url_for('sales'))


@app.route('/sales/<int:sid>/export/pdf')
def export_sale_pdf(sid):
    with get_db() as db:
        s = dict(db.execute("SELECT * FROM sales WHERE id=%s", (sid,)).fetchone())
        items = [dict(r) for r in db.execute("SELECT si.*, pr.name as product_name, pr.unit, pr.carton_qty FROM sale_items si JOIN products pr ON pr.id=si.product_id WHERE si.sale_id=%s", (sid,)).fetchall()]
        customer = None
        if s.get('customer_id'):
            row = db.execute("SELECT * FROM contacts WHERE id=%s", (s['customer_id'],)).fetchone()
            if row:
                customer = dict(row)
    company = get_settings()
    return send_file(io.BytesIO(generate_invoice_pdf(s, items, customer, company)), mimetype='application/pdf', download_name=f"Invoice-{s['num']}.pdf", as_attachment=True)


def _send_email(to_addr, subject, body, attachments=None, settings_dict=None):
    """Send an email via the configured SMTP server. Returns (ok, error_message)."""
    s = settings_dict or get_settings()
    host = s.get('smtp_host', '').strip()
    port = int(s.get('smtp_port') or 587)
    user = s.get('smtp_user', '').strip()
    pwd = s.get('smtp_pass', '')
    # IMPORTANT: From must usually match the authenticated user, or the server may silently drop/reject.
    # If smtp_from is set but differs from user, we prefer user to avoid spoofing issues.
    smtp_from_setting = (s.get('smtp_from') or '').strip()
    from_addr = smtp_from_setting if smtp_from_setting else (user or s.get('co_email') or '').strip()
    use_tls = bool(s.get('smtp_tls', '1'))
    if not host or not from_addr:
        return False, 'SMTP not configured — go to Settings → Email to set it up.'
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = from_addr
    msg['To'] = to_addr
    msg.set_content(body)
    for fname, fdata, mime in (attachments or []):
        maintype, subtype = mime.split('/', 1)
        msg.add_attachment(fdata, maintype=maintype, subtype=subtype, filename=fname)
    try:
        if port == 465:
            ctx = ssl.create_default_context()
            with smtplib.SMTP_SSL(host, port, context=ctx, timeout=20) as srv:
                if user: srv.login(user, pwd)
                srv.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=20) as srv:
                srv.ehlo()
                if use_tls:
                    srv.starttls(context=ssl.create_default_context())
                    srv.ehlo()
                if user: srv.login(user, pwd)
                srv.send_message(msg)
        return True, ''
    except Exception as e:
        return False, str(e)


@app.route('/settings/email/test', methods=['POST'])
@admin_required
def settings_email_test():
    to_addr = (request.form.get('to') or '').strip()
    if not to_addr:
        flash('Enter a test recipient email', 'error')
        return redirect(url_for('settings_page'))
    settings = get_settings()
    body = (
        f"This is a test email from your NEON installation.\n\n"
        f"If you received this in your inbox, your SMTP is correctly configured.\n"
        f"If it ended up in spam, check the From address matches the SMTP user, "
        f"and ask your hosting provider about SPF / DKIM / DMARC records for your domain.\n\n"
        f"SMTP host: {settings.get('smtp_host','')}\n"
        f"SMTP port: {settings.get('smtp_port','')}\n"
        f"From: {settings.get('smtp_from','')}\n"
    )
    ok, err = _send_email(to_addr,
                          subject='NEON SMTP test email',
                          body=body,
                          settings_dict=settings)
    if ok:
        flash(f'Test email sent to {to_addr}. Check your inbox AND spam folder.', 'success')
    else:
        flash(f'Test failed: {err}', 'error')
    return redirect(url_for('settings_page'))


@app.route('/sales/<int:sid>/email', methods=['POST'])
@admin_required
def email_invoice(sid):
    with get_db() as db:
        s_row = db.execute("SELECT * FROM sales WHERE id=%s", (sid,)).fetchone()
        if not s_row:
            flash('Invoice not found', 'error')
            return redirect(url_for('sales'))
        s = dict(s_row)
        items = [dict(r) for r in db.execute(
            "SELECT si.*, pr.name as product_name, pr.unit, pr.carton_qty FROM sale_items si JOIN products pr ON pr.id=si.product_id WHERE si.sale_id=%s",
            (sid,)).fetchall()]
        customer = None
        if s.get('customer_id'):
            row = db.execute("SELECT * FROM contacts WHERE id=%s", (s['customer_id'],)).fetchone()
            if row: customer = dict(row)
    company = get_settings()
    to_addr = (request.form.get('to') or '').strip()
    if not to_addr:
        to_addr = (customer or {}).get('email', '').strip()
    if not to_addr:
        flash('No recipient email — add one to the customer or enter it manually.', 'error')
        return redirect(url_for('view_sale', sid=sid))
    subject = request.form.get('subject') or f"Invoice {s.get('num', '')} from {company.get('co_company', 'us')}"
    body = request.form.get('body') or (
        f"Hello,\n\nPlease find attached your invoice {s.get('num', '')} dated {s.get('doc_date', '')}.\n\n"
        f"Total: £{s.get('total', 0):.2f}\n\nThank you for your business.\n\n"
        f"{company.get('co_company') or company.get('co_name') or ''}"
    )
    pdf_bytes = generate_invoice_pdf(s, items, customer, company)
    ok, err = _send_email(to_addr,
                          subject=subject,
                          body=body,
                          attachments=[(f"Invoice-{s.get('num','')}.pdf", pdf_bytes, 'application/pdf')],
                          settings_dict=company)
    if ok:
        flash(f'Invoice emailed to {to_addr}', 'success')
    else:
        flash(f'Failed to send email: {err}', 'error')
    return redirect(url_for('view_sale', sid=sid))


@app.route('/sales/<int:sid>/export/excel')
def export_sale_excel(sid):
    with get_db() as db:
        s = dict(db.execute("SELECT * FROM sales WHERE id=%s", (sid,)).fetchone())
        items = [dict(r) for r in db.execute("SELECT si.*, pr.name as product_name, pr.unit, pr.carton_qty FROM sale_items si JOIN products pr ON pr.id=si.product_id WHERE si.sale_id=%s", (sid,)).fetchall()]
        customer = None
        if s.get('customer_id'):
            row = db.execute("SELECT * FROM contacts WHERE id=%s", (s['customer_id'],)).fetchone()
            if row: customer = dict(row)
    company = get_settings()
    return send_file(io.BytesIO(generate_excel(s['num'], s, items, doc_type='sale', contact=customer, company=company)), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', download_name=f"{s['num']}.xlsx", as_attachment=True)


# ── Transfers ────────────────────────────────────────────────────────────────

@app.route('/transfers')
def transfers():
    with get_db() as db:
        rows = db.execute("""SELECT t.*, p.name as product_name, p.unit, w1.name as from_name, w2.name as to_name
            FROM transfers t JOIN products p ON p.id=t.product_id
            JOIN warehouses w1 ON w1.id=t.from_wh_id JOIN warehouses w2 ON w2.id=t.to_wh_id
            ORDER BY t.doc_date DESC""").fetchall()
    return render_template('transfers.html', transfers=rows)


@app.route('/transfers/add', methods=['GET', 'POST'])
@admin_required
def add_transfer():
    with get_db() as db:
        warehouses = db.execute("SELECT * FROM warehouses").fetchall()
        products = db.execute("SELECT * FROM products ORDER BY name").fetchall()
        if request.method == 'POST':
            pid = int(request.form['product_id'])
            from_id, to_id = int(request.form['from_wh_id']), int(request.form['to_wh_id'])
            qty = float(request.form['qty'])
            if from_id == to_id:
                flash('Source and destination must be different', 'error')
                return render_template('transfer_form.html', warehouses=warehouses, products=products, today=date.today().isoformat())
            avail = db.execute("SELECT qty FROM stock WHERE product_id=%s AND warehouse_id=%s", (pid, from_id)).fetchone()
            if not avail or avail['qty'] < qty:
                flash('Not enough stock in source warehouse', 'error')
                return render_template('transfer_form.html', warehouses=warehouses, products=products, today=date.today().isoformat())
            db.execute("UPDATE stock SET qty=qty-%s WHERE product_id=%s AND warehouse_id=%s", (qty, pid, from_id))
            db.execute("INSERT INTO stock(product_id,warehouse_id,qty) VALUES(%s,%s,%s) ON CONFLICT(product_id,warehouse_id) DO UPDATE SET qty=qty+excluded.qty", (pid, to_id, qty))
            db.execute("INSERT INTO transfers(doc_date,from_wh_id,to_wh_id,product_id,qty,notes) VALUES(%s,%s,%s,%s,%s,%s)",
                (request.form['doc_date'], from_id, to_id, pid, qty, request.form.get('notes','')))
            db.commit()
            flash('Transfer completed', 'success')
            return redirect(url_for('transfers'))
    return render_template('transfer_form.html', warehouses=warehouses, products=products, today=date.today().isoformat())


# ── Quotes ───────────────────────────────────────────────────────────────────

@app.route('/quotes')
def quotes():
    with get_db() as db:
        rows = db.execute("SELECT * FROM quotes ORDER BY doc_date DESC, id DESC").fetchall()
    return render_template('quotes.html', quotes=rows)


@app.route('/quotes/add', methods=['GET', 'POST'])
@admin_required
def add_quote():
    with get_db() as db:
        customers = [dict(r) for r in db.execute("SELECT * FROM contacts WHERE type='customer' ORDER BY name").fetchall()]
        warehouses = [dict(r) for r in db.execute("SELECT * FROM warehouses").fetchall()]
        products = [dict(r) for r in db.execute("SELECT p.*, COALESCE((SELECT SUM(qty) FROM stock WHERE product_id=p.id), 0) as stock_total FROM products p ORDER BY p.name").fetchall()]
        discounts = db.execute("SELECT * FROM discounts WHERE active=1").fetchall()
        if request.method == 'POST':
            global_disc = float(request.form.get('global_discount', 0))
            tax_pct = float(request.form.get('tax_pct', 0))
            pids = request.form.getlist('product_id[]')
            wids = request.form.getlist('warehouse_id[]')
            qtys = request.form.getlist('qty[]')
            prices = request.form.getlist('price[]')
            disc_pcts = request.form.getlist('discount_pct[]')
            subtotal = sum(float(q)*float(p)*(1-float(d)/100) for q,p,d in zip(qtys,prices,disc_pcts) if q and p) * (1 - global_disc/100)
            tax_amount = subtotal * tax_pct / 100
            total = subtotal + tax_amount
            num = next_num('QUO', 'quotes')
            customer_id = request.form.get('customer_id') or None
            cur = db.execute(
                "INSERT INTO quotes(num,doc_date,expiry_date,customer,customer_id,subtotal,total,discount,tax_pct,tax_amount,currency,status,notes) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (num, request.form['doc_date'], request.form.get('expiry_date',''),
                 request.form['customer'], customer_id, subtotal, total, global_disc,
                 tax_pct, tax_amount, request.form.get('currency','GBP'), 'draft', request.form.get('notes','')))
            qid = cur.lastrowid
            for pid, wid, qty, price, disc in zip(pids, wids, qtys, prices, disc_pcts):
                if pid and float(qty) > 0:
                    db.execute("INSERT INTO quote_items(quote_id,product_id,warehouse_id,qty,price,discount_pct) VALUES(%s,%s,%s,%s,%s,%s)",
                        (qid, int(pid), int(wid), float(qty), float(price), float(disc)))
            db.commit()
            flash(f'Quote {num} created', 'success')
            return redirect(url_for('view_quote', qid=qid))
    return render_template('quote_form.html', quote=None, items=None,
        customers=customers, warehouses=warehouses, products=products, discounts=discounts,
        today=date.today().isoformat())


@app.route('/quotes/<int:qid>')
def view_quote(qid):
    with get_db() as db:
        q = db.execute("SELECT * FROM quotes WHERE id=%s", (qid,)).fetchone()
        items = db.execute(
            "SELECT qi.*, pr.name as product_name, pr.unit, w.name as wh_name FROM quote_items qi JOIN products pr ON pr.id=qi.product_id JOIN warehouses w ON w.id=qi.warehouse_id WHERE qi.quote_id=%s",
            (qid,)).fetchall()
    return render_template('quote_view.html', quote=q, items=items)


@app.route('/quotes/<int:qid>/edit', methods=['GET', 'POST'])
@admin_required
def edit_quote(qid):
    with get_db() as db:
        quote = db.execute("SELECT * FROM quotes WHERE id=%s", (qid,)).fetchone()
        if not quote:
            flash('Quote not found', 'error')
            return redirect(url_for('quotes'))
        customers = [dict(r) for r in db.execute("SELECT * FROM contacts WHERE type='customer' ORDER BY name").fetchall()]
        warehouses = [dict(r) for r in db.execute("SELECT * FROM warehouses").fetchall()]
        products = [dict(r) for r in db.execute("SELECT p.*, COALESCE((SELECT SUM(qty) FROM stock WHERE product_id=p.id), 0) as stock_total FROM products p ORDER BY p.name").fetchall()]
        discounts = db.execute("SELECT * FROM discounts WHERE active=1").fetchall()
        items = [dict(r) for r in db.execute("SELECT * FROM quote_items WHERE quote_id=%s", (qid,)).fetchall()]
        if request.method == 'POST':
            pids = request.form.getlist('product_id[]')
            wids = request.form.getlist('warehouse_id[]')
            qtys = request.form.getlist('qty[]')
            prices = request.form.getlist('price[]')
            disc_pcts = request.form.getlist('discount_pct[]')
            global_disc = float(request.form.get('global_discount', 0))
            tax_pct = float(request.form.get('tax_pct', 0))
            subtotal = sum(float(q)*float(p)*(1-float(d)/100) for q,p,d in zip(qtys,prices,disc_pcts) if q and p) * (1 - global_disc/100)
            tax_amount = subtotal * tax_pct / 100
            total = subtotal + tax_amount
            customer_id = request.form.get('customer_id') or None
            db.execute("UPDATE quotes SET doc_date=%s,expiry_date=%s,customer=%s,customer_id=%s,subtotal=%s,total=%s,discount=%s,tax_pct=%s,tax_amount=%s,currency=%s,notes=%s WHERE id=%s",
                (request.form['doc_date'], request.form.get('expiry_date',''),
                 request.form['customer'], customer_id, subtotal, total, global_disc,
                 tax_pct, tax_amount, request.form.get('currency','GBP'), request.form.get('notes',''), qid))
            db.execute("DELETE FROM quote_items WHERE quote_id=%s", (qid,))
            for pid, wid, qty, price, disc in zip(pids, wids, qtys, prices, disc_pcts):
                if pid and float(qty) > 0:
                    db.execute("INSERT INTO quote_items(quote_id,product_id,warehouse_id,qty,price,discount_pct) VALUES(%s,%s,%s,%s,%s,%s)",
                        (qid, int(pid), int(wid), float(qty), float(price), float(disc)))
            db.commit()
            flash(f'Quote {quote["num"]} updated', 'success')
            return redirect(url_for('view_quote', qid=qid))
    return render_template('quote_form.html', quote=dict(quote), items=items,
        customers=customers, warehouses=warehouses, products=products, discounts=discounts)


@app.route('/quotes/<int:qid>/status', methods=['POST'])
@admin_required
def update_quote_status(qid):
    status = request.form.get('status')
    if status in ('draft', 'sent', 'accepted', 'rejected', 'expired'):
        with get_db() as db:
            db.execute("UPDATE quotes SET status=%s WHERE id=%s", (status, qid))
            db.commit()
    return redirect(url_for('view_quote', qid=qid))


@app.route('/quotes/<int:qid>/convert', methods=['POST'])
@admin_required
def convert_quote(qid):
    with get_db() as db:
        q = db.execute("SELECT * FROM quotes WHERE id=%s", (qid,)).fetchone()
        items = db.execute("SELECT * FROM quote_items WHERE quote_id=%s", (qid,)).fetchall()
        errors = []
        for item in items:
            avail = db.execute("SELECT qty FROM stock WHERE product_id=%s AND warehouse_id=%s", (item['product_id'], item['warehouse_id'])).fetchone()
            available = avail['qty'] if avail else 0
            if available < item['qty']:
                pname = db.execute("SELECT name FROM products WHERE id=%s", (item['product_id'],)).fetchone()['name']
                errors.append(f'Not enough stock for {pname} (available: {available:.2f})')
        if errors:
            for e in errors: flash(e, 'error')
            return redirect(url_for('view_quote', qid=qid))
        num = next_num('SAL', 'sales')
        cur = db.execute(
            "INSERT INTO sales(num,doc_date,customer,customer_id,subtotal,total,discount,tax_pct,tax_amount,currency,status,notes) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (num, date.today().isoformat(), q['customer'], q['customer_id'], q['subtotal'], q['total'],
             q['discount'], q['tax_pct'], q['tax_amount'], q['currency'], 'completed', f'Converted from {q["num"]}'))
        sale_id = cur.lastrowid
        for item in items:
            db.execute("INSERT INTO sale_items(sale_id,product_id,warehouse_id,qty,price,discount_pct) VALUES(%s,%s,%s,%s,%s,%s)",
                (sale_id, item['product_id'], item['warehouse_id'], item['qty'], item['price'], item['discount_pct']))
            db.execute("UPDATE stock SET qty=qty-%s WHERE product_id=%s AND warehouse_id=%s", (item['qty'], item['product_id'], item['warehouse_id']))
        db.execute("UPDATE quotes SET status='accepted' WHERE id=%s", (qid,))
        db.commit()
    flash(f'Quote converted to sale {num}', 'success')
    return redirect(url_for('view_sale', sid=sale_id))


@app.route('/quotes/<int:qid>/delete', methods=['POST'])
@admin_required
def delete_quote(qid):
    with get_db() as db:
        db.execute("DELETE FROM quotes WHERE id=%s", (qid,))
        db.commit()
    flash('Quote deleted', 'success')
    return redirect(url_for('quotes'))


@app.route('/quotes/bulk', methods=['POST'])
@admin_required
def bulk_quotes():
    ids = request.form.getlist('ids[]')
    action = request.form.get('action')
    if not ids:
        return redirect(url_for('quotes'))
    count = 0
    with get_db() as db:
        for qid in ids:
            if action == 'delete':
                db.execute("DELETE FROM quotes WHERE id=%s", (qid,))
            elif action in ('draft', 'sent', 'accepted', 'rejected', 'expired'):
                db.execute("UPDATE quotes SET status=%s WHERE id=%s", (action, qid))
            count += 1
        db.commit()
    flash(f'{count} quote(s) updated', 'success')
    return redirect(url_for('quotes'))


@app.route('/quotes/<int:qid>/export/pdf')
def export_quote_pdf(qid):
    with get_db() as db:
        q = dict(db.execute("SELECT * FROM quotes WHERE id=%s", (qid,)).fetchone())
        items = [dict(r) for r in db.execute(
            "SELECT qi.*, pr.name as product_name, pr.unit FROM quote_items qi JOIN products pr ON pr.id=qi.product_id WHERE qi.quote_id=%s",
            (qid,)).fetchall()]
        customer = None
        if q.get('customer_id'):
            row = db.execute("SELECT * FROM contacts WHERE id=%s", (q['customer_id'],)).fetchone()
            if row:
                customer = dict(row)
        company = get_settings()
    return send_file(io.BytesIO(generate_invoice_pdf(q, items, customer, company, doc_title='QUOTE')),
        mimetype='application/pdf', download_name=f"{q['num']}.pdf", as_attachment=True)


# ── TTNs ─────────────────────────────────────────────────────────────────────

@app.route('/ttns')
def ttns():
    with get_db() as db:
        rows = db.execute("SELECT * FROM ttns ORDER BY doc_date DESC").fetchall()
    return render_template('ttns.html', ttns=rows)


@app.route('/ttns/add', methods=['GET', 'POST'])
@admin_required
def add_ttn():
    if request.method == 'POST':
        with get_db() as db:
            num = next_num('TTN', 'ttns')
            db.execute("INSERT INTO ttns(num,doc_date,sender,receiver,carrier,ref_doc,status,notes) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)",
                (num, request.form['doc_date'], request.form.get('sender',''), request.form.get('receiver',''),
                 request.form.get('carrier',''), request.form.get('ref_doc',''),
                 request.form.get('status','in_transit'), request.form.get('notes','')))
            db.commit()
        flash(f'TTN {num} created', 'success')
        return redirect(url_for('ttns'))
    return render_template('ttn_form.html', ttn=None, today=date.today().isoformat())


@app.route('/ttns/edit/<int:tid>', methods=['GET', 'POST'])
@admin_required
def edit_ttn(tid):
    with get_db() as db:
        ttn = db.execute("SELECT * FROM ttns WHERE id=%s", (tid,)).fetchone()
        if request.method == 'POST':
            db.execute("UPDATE ttns SET doc_date=%s,sender=%s,receiver=%s,carrier=%s,ref_doc=%s,status=%s,notes=%s WHERE id=%s",
                (request.form['doc_date'], request.form.get('sender',''), request.form.get('receiver',''),
                 request.form.get('carrier',''), request.form.get('ref_doc',''),
                 request.form.get('status','in_transit'), request.form.get('notes',''), tid))
            db.commit()
            flash('TTN updated', 'success')
            return redirect(url_for('ttns'))
    return render_template('ttn_form.html', ttn=ttn, today=date.today().isoformat())


# ── Finance ──────────────────────────────────────────────────────────────────

@app.route('/finance')
def finance():
    method   = request.args.get('method', '')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    with get_db() as db:
        base = "SELECT * FROM transactions WHERE 1=1"
        params = []
        if method:     base += " AND method=%s";    params.append(method)
        if date_from:  base += " AND doc_date>=%s"; params.append(date_from)
        if date_to:    base += " AND doc_date<=%s"; params.append(date_to)
        rows = db.execute(base + " ORDER BY doc_date DESC", params).fetchall()
        # Summary metrics respect the same filters
        sum_base = "SELECT COALESCE(SUM(amount),0) AS s FROM transactions WHERE 1=1"
        sp = []
        if method:    sum_base += " AND method=%s";    sp.append(method)
        if date_from: sum_base += " AND doc_date>=%s"; sp.append(date_from)
        if date_to:   sum_base += " AND doc_date<=%s"; sp.append(date_to)
        income   = db.execute(sum_base + " AND type='income'",  sp).fetchone()['s']
        expenses = db.execute(sum_base + " AND type='expense'", sp).fetchone()['s']
        cash_in  = db.execute(sum_base + " AND type='income'  AND method='cash'", sp).fetchone()['s']
        cash_out = db.execute(sum_base + " AND type='expense' AND method='cash'", sp).fetchone()['s']
        bank_in  = db.execute(sum_base + " AND type='income'  AND method='bank'", sp).fetchone()['s']
        bank_out = db.execute(sum_base + " AND type='expense' AND method='bank'", sp).fetchone()['s']
    return render_template('finance.html', transactions=rows, income=income, expenses=expenses,
        net=income-expenses, cash_balance=cash_in-cash_out, bank_balance=bank_in-bank_out,
        method=method, date_from=date_from, date_to=date_to)


@app.route('/finance/add', methods=['GET', 'POST'])
@admin_required
def add_transaction():
    with get_db() as db:
        contacts = db.execute("SELECT id, name FROM contacts ORDER BY name").fetchall()
        if request.method == 'POST':
            contact_id = request.form.get('contact_id') or None
            contact_name = ''
            if contact_id:
                c = db.execute("SELECT name FROM contacts WHERE id=%s", (contact_id,)).fetchone()
                contact_name = c['name'] if c else ''
            db.execute('INSERT INTO transactions(doc_date,type,method,"desc",amount,currency,contact_id,contact,ref_doc) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (request.form['doc_date'], request.form['type'], request.form.get('method','cash'),
                 request.form['desc'], float(request.form.get('amount', 0)),
                 request.form.get('currency','GBP'), contact_id, contact_name, request.form.get('ref_doc','')))
            db.commit()
            flash('Transaction added', 'success')
            return redirect(url_for('finance'))
    return render_template('transaction_form.html', contacts=contacts, today=date.today().isoformat())


@app.route('/finance/edit/<int:tid>', methods=['GET', 'POST'])
@admin_required
def edit_transaction(tid):
    with get_db() as db:
        t = db.execute("SELECT * FROM transactions WHERE id=%s", (tid,)).fetchone()
        contacts = db.execute("SELECT id, name FROM contacts ORDER BY name").fetchall()
        if request.method == 'POST':
            contact_id = request.form.get('contact_id') or None
            contact_name = ''
            if contact_id:
                c = db.execute("SELECT name FROM contacts WHERE id=%s", (contact_id,)).fetchone()
                contact_name = c['name'] if c else ''
            db.execute('UPDATE transactions SET doc_date=%s,type=%s,method=%s,"desc"=%s,amount=%s,currency=%s,contact_id=%s,contact=%s,ref_doc=%s WHERE id=%s',
                (request.form['doc_date'], request.form['type'], request.form.get('method','cash'),
                 request.form['desc'], float(request.form.get('amount', 0)),
                 request.form.get('currency','GBP'), contact_id, contact_name, request.form.get('ref_doc',''), tid))
            db.commit()
            flash('Transaction updated', 'success')
            return redirect(url_for('finance'))
    return render_template('transaction_form.html', transaction=t, contacts=contacts, today=date.today().isoformat())


@app.route('/finance/delete/<int:tid>', methods=['POST'])
@admin_required
def delete_transaction(tid):
    with get_db() as db:
        db.execute("DELETE FROM transactions WHERE id=%s", (tid,))
        db.commit()
    flash('Transaction deleted', 'success')
    return redirect(url_for('finance'))


# ── Exchange Rates ────────────────────────────────────────────────────────────

# ── Discounts ─────────────────────────────────────────────────────────────────

@app.route('/discounts')
def discounts():
    with get_db() as db:
        rows = db.execute("SELECT * FROM discounts ORDER BY name").fetchall()
    return render_template('discounts.html', discounts=rows)


@app.route('/discounts/add', methods=['GET', 'POST'])
@admin_required
def add_discount():
    if request.method == 'POST':
        with get_db() as db:
            db.execute("INSERT INTO discounts(name,type,value,applies_to,active) VALUES(%s,%s,%s,%s,1)",
                (request.form['name'], request.form['type'], float(request.form['value']), request.form.get('applies_to','all')))
            db.commit()
        flash('Discount added', 'success')
        return redirect(url_for('discounts'))
    return render_template('discount_form.html', discount=None)


@app.route('/discounts/toggle/<int:did>', methods=['POST'])
@admin_required
def toggle_discount(did):
    with get_db() as db:
        db.execute("UPDATE discounts SET active=1-active WHERE id=%s", (did,))
        db.commit()
    return redirect(url_for('discounts'))


# ── Reports ───────────────────────────────────────────────────────────────────

@app.route('/reports')
def reports():
    return render_template('reports.html')


@app.route('/reports/sales')
def report_sales():
    date_from, date_to = request.args.get('from', ''), request.args.get('to', '')
    with get_db() as db:
        q = """SELECT s.num, s.doc_date, s.customer, si.product_id, si.qty, si.price, si.discount_pct,
               p.name as product_name, p.cost FROM sales s
               JOIN sale_items si ON si.sale_id=s.id JOIN products p ON p.id=si.product_id
               WHERE s.status='completed'"""
        params = []
        if date_from: q += " AND s.doc_date>=%s"; params.append(date_from)
        if date_to: q += " AND s.doc_date<=%s"; params.append(date_to)
        rows = db.execute(q + " ORDER BY s.doc_date DESC", params).fetchall()
        by_product = {}
        for r in rows:
            pid = r['product_id']
            if pid not in by_product:
                by_product[pid] = {'name': r['product_name'], 'qty': 0, 'revenue': 0, 'cost': 0, 'profit': 0}
            rev = r['qty'] * r['price'] * (1 - r['discount_pct']/100)
            cost = r['qty'] * r['cost']
            by_product[pid]['qty'] += r['qty']
            by_product[pid]['revenue'] += rev
            by_product[pid]['cost'] += cost
            by_product[pid]['profit'] += rev - cost
        by_customer = {}
        for r in rows:
            c = r['customer']
            if c not in by_customer:
                by_customer[c] = {'revenue': 0, 'orders': set()}
            by_customer[c]['revenue'] += r['qty'] * r['price'] * (1 - r['discount_pct']/100)
            by_customer[c]['orders'].add(r['num'])
        for c in by_customer:
            by_customer[c]['orders'] = len(by_customer[c]['orders'])
        total_revenue = sum(v['revenue'] for v in by_product.values())
        total_profit = sum(v['profit'] for v in by_product.values())
    return render_template('report_sales.html',
        by_product=sorted(by_product.values(), key=lambda x: x['revenue'], reverse=True),
        by_customer=sorted(by_customer.items(), key=lambda x: x[1]['revenue'], reverse=True),
        total_revenue=total_revenue, total_profit=total_profit, date_from=date_from, date_to=date_to)


@app.route('/reports/stock')
def report_stock():
    with get_db() as db:
        products = db.execute("SELECT * FROM products ORDER BY category, name").fetchall()
        warehouses = db.execute("SELECT * FROM warehouses").fetchall()
        rows, total_value = [], 0
        for p in products:
            stock_by_wh, total = {}, 0
            for w in warehouses:
                r = db.execute("SELECT qty FROM stock WHERE product_id=%s AND warehouse_id=%s", (p['id'], w['id'])).fetchone()
                q = r['qty'] if r else 0
                stock_by_wh[w['id']] = q
                total += q
            value = total * p['cost']
            total_value += value
            sold = db.execute("SELECT COALESCE(SUM(qty),0) AS s FROM sale_items WHERE product_id=%s", (p['id'],)).fetchone()['s']
            rows.append({'product': p, 'stock': stock_by_wh, 'total': total, 'value': value, 'sold': sold})
    return render_template('report_stock.html', rows=rows, warehouses=warehouses, total_value=total_value)


@app.route('/reports/turnover')
def report_turnover():
    with get_db() as db:
        rows = []
        for p in db.execute("SELECT * FROM products ORDER BY name").fetchall():
            in_qty = db.execute("SELECT COALESCE(SUM(qty),0) AS s FROM purchase_items WHERE product_id=%s", (p['id'],)).fetchone()['s']
            out_qty = db.execute("SELECT COALESCE(SUM(qty),0) AS s FROM sale_items WHERE product_id=%s", (p['id'],)).fetchone()['s']
            wo_qty = db.execute("SELECT COALESCE(SUM(qty),0) AS s FROM writeoff_items WHERE product_id=%s", (p['id'],)).fetchone()['s']
            curr = db.execute("SELECT COALESCE(SUM(qty),0) AS s FROM stock WHERE product_id=%s", (p['id'],)).fetchone()['s']
            rows.append({'name': p['name'], 'sku': p['sku'], 'unit': p['unit'],
                         'in': in_qty, 'out': out_qty, 'writeoff': wo_qty, 'stock': curr,
                         'turnover': round(out_qty / max(curr, 1), 2)})
    return render_template('report_turnover.html', rows=rows)


@app.route('/reports/profit')
def report_profit():
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    customer_filter = request.args.get('customer', '')
    with get_db() as db:
        customers = [r[0] for r in db.execute(
            "SELECT DISTINCT customer FROM sales WHERE status='completed' AND customer IS NOT NULL AND customer!='' ORDER BY customer"
        ).fetchall()]
        q = """SELECT s.doc_date, s.num, s.customer,
               SUM(si.qty*si.price*(1-si.discount_pct/100)) as revenue,
               SUM(si.qty*p.cost) as cost
               FROM sales s JOIN sale_items si ON si.sale_id=s.id
               JOIN products p ON p.id=si.product_id WHERE s.status='completed'"""
        params = []
        if date_from: q += " AND s.doc_date>=%s"; params.append(date_from)
        if date_to: q += " AND s.doc_date<=%s"; params.append(date_to)
        if customer_filter: q += " AND s.customer=%s"; params.append(customer_filter)
        rows = db.execute(q + " GROUP BY s.id, s.doc_date, s.num, s.customer ORDER BY s.doc_date DESC", params).fetchall()
        total_rev = sum(r['revenue'] for r in rows)
        total_cost = sum(r['cost'] for r in rows)
    return render_template('report_profit.html', rows=rows, total_rev=total_rev,
        total_cost=total_cost, total_profit=total_rev-total_cost,
        date_from=date_from, date_to=date_to,
        customers=customers, customer_filter=customer_filter)


@app.route('/reports/receivables')
def report_receivables():
    today = date.today()
    with get_db() as db:
        sales = db.execute("""SELECT s.id, s.num, s.doc_date, s.customer, s.total, s.customer_id
            FROM sales s WHERE s.status='completed' AND s.paid=0 AND s.archived=0
            ORDER BY s.doc_date""").fetchall()
        rows = []
        buckets = {'current': 0, '30': 0, '60': 0, '90': 0, '90+': 0}
        for s in sales:
            try:
                d = datetime.strptime(s['doc_date'], '%Y-%m-%d').date()
                age = (today - d).days
            except Exception:
                age = 0
            if   age <=  0: bucket = 'current'
            elif age <= 30: bucket = '30'
            elif age <= 60: bucket = '60'
            elif age <= 90: bucket = '90'
            else:           bucket = '90+'
            buckets[bucket] += s['total']
            rows.append({'sale': dict(s), 'age': age, 'bucket': bucket})
        total = sum(s['total'] for s in sales)
    return render_template('report_receivables.html', rows=rows, buckets=buckets, total=total)


@app.route('/reports/cashflow')
def report_cashflow():
    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    with get_db() as db:
        q = "SELECT doc_date, type, amount FROM transactions WHERE 1=1"
        params = []
        if date_from: q += " AND doc_date>=%s"; params.append(date_from)
        if date_to:   q += " AND doc_date<=%s"; params.append(date_to)
        rows = db.execute(q + " ORDER BY doc_date", params).fetchall()
        by_month = {}
        for r in rows:
            month = r['doc_date'][:7] if r['doc_date'] else ''
            if month not in by_month:
                by_month[month] = {'income': 0, 'expense': 0}
            by_month[month][r['type']] = by_month[month].get(r['type'], 0) + r['amount']
        months = sorted(by_month.items(), reverse=True)
        total_income  = sum(m['income']  for _, m in months)
        total_expense = sum(m['expense'] for _, m in months)
    return render_template('report_cashflow.html', months=months,
        total_income=total_income, total_expense=total_expense,
        date_from=date_from, date_to=date_to)


@app.route('/reports/low-stock')
def report_low_stock():
    with get_db() as db:
        products = db.execute("SELECT * FROM products WHERE min_stock > 0 ORDER BY name").fetchall()
        rows = []
        for p in products:
            total = db.execute("SELECT COALESCE(SUM(qty),0) AS s FROM stock WHERE product_id=%s", (p['id'],)).fetchone()['s']
            if total <= p['min_stock']:
                short = max(0, p['min_stock'] - total)
                rows.append({'product': dict(p), 'stock': total, 'shortage': short})
    return render_template('report_low_stock.html', rows=rows)


@app.route('/reports/top-customers')
def report_top_customers():
    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    with get_db() as db:
        q = """SELECT s.id, s.customer, s.total, s.paid,
               COALESCE(SUM(si.qty*si.price*(1-si.discount_pct/100)), 0) as gross_revenue,
               COALESCE(SUM(si.qty*p.cost), 0) as cogs
               FROM sales s
               LEFT JOIN sale_items si ON si.sale_id=s.id
               LEFT JOIN products p ON p.id=si.product_id
               WHERE s.status='completed' AND s.archived=0"""
        params = []
        if date_from: q += " AND s.doc_date>=%s"; params.append(date_from)
        if date_to:   q += " AND s.doc_date<=%s"; params.append(date_to)
        q += " GROUP BY s.id, s.customer, s.total, s.paid"
        rows = db.execute(q, params).fetchall()
        agg = {}
        for r in rows:
            c = r['customer'] or 'Unknown'
            if c not in agg:
                agg[c] = {'name': c, 'orders': 0, 'revenue': 0, 'cost': 0, 'profit': 0, 'paid': 0, 'unpaid': 0}
            agg[c]['orders']  += 1
            agg[c]['revenue'] += r['total']
            agg[c]['cost']    += r['cogs']
            agg[c]['profit']  += r['gross_revenue'] - r['cogs']
            if r['paid']: agg[c]['paid']   += r['total']
            else:         agg[c]['unpaid'] += r['total']
        for c in agg.values():
            c['margin_pct'] = (c['profit'] / c['revenue'] * 100) if c['revenue'] else 0
        customers     = sorted(agg.values(), key=lambda x: x['revenue'], reverse=True)
        total_revenue = sum(c['revenue'] for c in customers)
        total_cost    = sum(c['cost']    for c in customers)
        total_profit  = sum(c['profit']  for c in customers)
    return render_template('report_top_customers.html', customers=customers,
        total_revenue=total_revenue, total_cost=total_cost, total_profit=total_profit,
        date_from=date_from, date_to=date_to)


@app.route('/reports/vat')
def report_vat():
    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    with get_db() as db:
        q = "SELECT doc_date, total, tax_amount, tax_pct FROM sales WHERE status='completed' AND archived=0 AND tax_amount>0"
        params = []
        if date_from: q += " AND doc_date>=%s"; params.append(date_from)
        if date_to:   q += " AND doc_date<=%s"; params.append(date_to)
        sales_rows = db.execute(q + " ORDER BY doc_date", params).fetchall()
        by_month = {}
        for r in sales_rows:
            month = r['doc_date'][:7] if r['doc_date'] else ''
            if month not in by_month:
                by_month[month] = {'net': 0, 'vat': 0, 'gross': 0}
            tax = r['tax_amount'] or 0
            net = (r['total'] or 0) - tax
            by_month[month]['net']   += net
            by_month[month]['vat']   += tax
            by_month[month]['gross'] += r['total'] or 0
        months = sorted(by_month.items(), reverse=True)
        total_net   = sum(m['net']   for _, m in months)
        total_vat   = sum(m['vat']   for _, m in months)
        total_gross = sum(m['gross'] for _, m in months)
    return render_template('report_vat.html', months=months,
        total_net=total_net, total_vat=total_vat, total_gross=total_gross,
        date_from=date_from, date_to=date_to)


# ── Credit Notes ──────────────────────────────────────────────────────────────

@app.route('/credit-notes')
def credit_notes():
    with get_db() as db:
        rows = db.execute("""
            SELECT cn.*, wh.name as wh_name
            FROM credit_notes cn
            LEFT JOIN warehouses wh ON wh.id = cn.warehouse_id
            ORDER BY cn.doc_date DESC, cn.id DESC
        """).fetchall()
    return render_template('credit_notes.html', notes=rows)


@app.route('/credit-notes/add', methods=['GET', 'POST'])
@admin_required
def add_credit_note():
    with get_db() as db:
        customers = db.execute("SELECT * FROM contacts WHERE type='customer' ORDER BY name").fetchall()
        warehouses = db.execute("SELECT * FROM warehouses").fetchall()
        products = [dict(r) for r in db.execute("SELECT * FROM products ORDER BY name").fetchall()]
        sale_id = request.args.get('sale_id', type=int)
        prefill_sale = None
        prefill_items = []
        if sale_id:
            prefill_sale = db.execute("SELECT * FROM sales WHERE id=%s", (sale_id,)).fetchone()
            prefill_items = db.execute(
                "SELECT si.*, p.name as product_name, p.unit FROM sale_items si JOIN products p ON p.id=si.product_id WHERE si.sale_id=%s",
                (sale_id,)).fetchall()
        if request.method == 'POST':
            pids   = request.form.getlist('product_id[]')
            qtys   = request.form.getlist('qty[]')
            prices = request.form.getlist('price[]')
            items_data = [(int(p), float(q), float(pr)) for p, q, pr in zip(pids, qtys, prices) if p and float(q) > 0]
            total = sum(q * pr for _, q, pr in items_data)
            num = next_num('CN', 'credit_notes')
            wid = request.form.get('warehouse_id') or None
            customer_id = request.form.get('customer_id') or None
            cur = db.execute(
                "INSERT INTO credit_notes(num,doc_date,sale_id,customer,customer_id,warehouse_id,total,status,notes,created_at) VALUES(%s,%s,%s,%s,%s,%s,%s,'draft',%s,%s)",
                (num, request.form['doc_date'],
                 request.form.get('sale_id') or None,
                 request.form['customer'], customer_id, wid, total,
                 request.form.get('notes', ''),
                 datetime.utcnow().isoformat()))
            cn_id = cur.lastrowid
            for pid, qty, price in items_data:
                pname = db.execute("SELECT name FROM products WHERE id=%s", (pid,)).fetchone()['name']
                db.execute("INSERT INTO credit_note_items(credit_note_id,product_id,qty,price,product_name) VALUES(%s,%s,%s,%s,%s)",
                           (cn_id, pid, qty, price, pname))
            db.commit()
            flash(f'Credit note {num} created', 'success')
            return redirect(url_for('view_credit_note', cnid=cn_id))
    return render_template('credit_note_form.html',
        customers=customers, warehouses=warehouses, products=products,
        today=date.today().isoformat(), prefill_sale=prefill_sale, prefill_items=prefill_items)


@app.route('/credit-notes/<int:cnid>')
def view_credit_note(cnid):
    with get_db() as db:
        cn = db.execute("""
            SELECT cn.*, wh.name as wh_name
            FROM credit_notes cn LEFT JOIN warehouses wh ON wh.id=cn.warehouse_id
            WHERE cn.id=%s""", (cnid,)).fetchone()
        items = db.execute("""
            SELECT cni.*, p.unit
            FROM credit_note_items cni
            LEFT JOIN products p ON p.id=cni.product_id
            WHERE cni.credit_note_id=%s""", (cnid,)).fetchall()
    return render_template('credit_note_view.html', cn=cn, items=items, co=get_settings())


@app.route('/credit-notes/<int:cnid>/apply', methods=['POST'])
@admin_required
def apply_credit_note(cnid):
    with get_db() as db:
        cn = db.execute("SELECT * FROM credit_notes WHERE id=%s", (cnid,)).fetchone()
        if cn['status'] == 'applied':
            flash('Credit note already applied', 'error')
            return redirect(url_for('view_credit_note', cnid=cnid))
        items = db.execute("SELECT * FROM credit_note_items WHERE credit_note_id=%s", (cnid,)).fetchall()
        if cn['warehouse_id']:
            for item in items:
                db.execute("INSERT INTO stock(product_id,warehouse_id,qty) VALUES(%s,%s,%s) ON CONFLICT(product_id,warehouse_id) DO UPDATE SET qty=qty+excluded.qty",
                           (item['product_id'], cn['warehouse_id'], item['qty']))
        if cn['customer_id']:
            db.execute("UPDATE contacts SET balance=balance-%s WHERE id=%s", (cn['total'], cn['customer_id']))
        db.execute("UPDATE credit_notes SET status='applied' WHERE id=%s", (cnid,))
        db.commit()
    flash('Credit note applied — stock restored and customer balance updated', 'success')
    return redirect(url_for('view_credit_note', cnid=cnid))


@app.route('/credit-notes/<int:cnid>/delete', methods=['POST'])
@admin_required
def delete_credit_note(cnid):
    with get_db() as db:
        cn = db.execute("SELECT * FROM credit_notes WHERE id=%s", (cnid,)).fetchone()
        if cn['status'] == 'applied':
            flash('Cannot delete an applied credit note', 'error')
            return redirect(url_for('view_credit_note', cnid=cnid))
        db.execute("DELETE FROM credit_note_items WHERE credit_note_id=%s", (cnid,))
        db.execute("DELETE FROM credit_notes WHERE id=%s", (cnid,))
        db.commit()
    flash('Credit note deleted', 'success')
    return redirect(url_for('credit_notes'))


# ── Debit Notes ───────────────────────────────────────────────────────────────

@app.route('/debit-notes')
def debit_notes():
    with get_db() as db:
        rows = db.execute("""
            SELECT dn.*, wh.name as wh_name
            FROM debit_notes dn
            LEFT JOIN warehouses wh ON wh.id = dn.warehouse_id
            ORDER BY dn.doc_date DESC, dn.id DESC
        """).fetchall()
    return render_template('debit_notes.html', notes=rows)


@app.route('/debit-notes/add', methods=['GET', 'POST'])
@admin_required
def add_debit_note():
    with get_db() as db:
        suppliers = db.execute("SELECT * FROM contacts WHERE type='supplier' ORDER BY name").fetchall()
        warehouses = db.execute("SELECT * FROM warehouses").fetchall()
        products = [dict(r) for r in db.execute("SELECT * FROM products ORDER BY name").fetchall()]
        purchase_id = request.args.get('purchase_id', type=int)
        prefill_purchase = None
        prefill_items = []
        if purchase_id:
            prefill_purchase = db.execute("SELECT * FROM purchases WHERE id=%s", (purchase_id,)).fetchone()
            prefill_items = db.execute(
                "SELECT pi.*, p.name as product_name, p.unit FROM purchase_items pi JOIN products p ON p.id=pi.product_id WHERE pi.purchase_id=%s",
                (purchase_id,)).fetchall()
        if request.method == 'POST':
            pids   = request.form.getlist('product_id[]')
            qtys   = request.form.getlist('qty[]')
            prices = request.form.getlist('price[]')
            items_data = [(int(p), float(q), float(pr)) for p, q, pr in zip(pids, qtys, prices) if p and float(q) > 0]
            total = sum(q * pr for _, q, pr in items_data)
            num = next_num('DN', 'debit_notes')
            wid = request.form.get('warehouse_id') or None
            supplier_id = request.form.get('supplier_id') or None
            cur = db.execute(
                "INSERT INTO debit_notes(num,doc_date,purchase_id,supplier,supplier_id,warehouse_id,total,status,notes,created_at) VALUES(%s,%s,%s,%s,%s,%s,%s,'draft',%s,%s)",
                (num, request.form['doc_date'],
                 request.form.get('purchase_id') or None,
                 request.form['supplier'], supplier_id, wid, total,
                 request.form.get('notes', ''),
                 datetime.utcnow().isoformat()))
            dn_id = cur.lastrowid
            for pid, qty, price in items_data:
                pname = db.execute("SELECT name FROM products WHERE id=%s", (pid,)).fetchone()['name']
                db.execute("INSERT INTO debit_note_items(debit_note_id,product_id,qty,price,product_name) VALUES(%s,%s,%s,%s,%s)",
                           (dn_id, pid, qty, price, pname))
            db.commit()
            flash(f'Debit note {num} created', 'success')
            return redirect(url_for('view_debit_note', dnid=dn_id))
    return render_template('debit_note_form.html',
        suppliers=suppliers, warehouses=warehouses, products=products,
        today=date.today().isoformat(), prefill_purchase=prefill_purchase, prefill_items=prefill_items)


@app.route('/debit-notes/<int:dnid>')
def view_debit_note(dnid):
    with get_db() as db:
        dn = db.execute("""
            SELECT dn.*, wh.name as wh_name
            FROM debit_notes dn LEFT JOIN warehouses wh ON wh.id=dn.warehouse_id
            WHERE dn.id=%s""", (dnid,)).fetchone()
        items = db.execute("""
            SELECT dni.*, p.unit
            FROM debit_note_items dni
            LEFT JOIN products p ON p.id=dni.product_id
            WHERE dni.debit_note_id=%s""", (dnid,)).fetchall()
    return render_template('debit_note_view.html', dn=dn, items=items)


@app.route('/debit-notes/<int:dnid>/apply', methods=['POST'])
@admin_required
def apply_debit_note(dnid):
    with get_db() as db:
        dn = db.execute("SELECT * FROM debit_notes WHERE id=%s", (dnid,)).fetchone()
        if dn['status'] == 'applied':
            flash('Debit note already applied', 'error')
            return redirect(url_for('view_debit_note', dnid=dnid))
        items = db.execute("SELECT * FROM debit_note_items WHERE debit_note_id=%s", (dnid,)).fetchall()
        if dn['warehouse_id']:
            for item in items:
                db.execute("UPDATE stock SET qty=qty-%s WHERE product_id=%s AND warehouse_id=%s",
                           (item['qty'], item['product_id'], dn['warehouse_id']))
        if dn['supplier_id']:
            db.execute("UPDATE contacts SET balance=balance+%s WHERE id=%s", (dn['total'], dn['supplier_id']))
        db.execute("UPDATE debit_notes SET status='applied' WHERE id=%s", (dnid,))
        db.commit()
    flash('Debit note applied — stock removed and supplier balance updated', 'success')
    return redirect(url_for('view_debit_note', dnid=dnid))


@app.route('/debit-notes/<int:dnid>/delete', methods=['POST'])
@admin_required
def delete_debit_note(dnid):
    with get_db() as db:
        dn = db.execute("SELECT * FROM debit_notes WHERE id=%s", (dnid,)).fetchone()
        if dn['status'] == 'applied':
            flash('Cannot delete an applied debit note', 'error')
            return redirect(url_for('view_debit_note', dnid=dnid))
        db.execute("DELETE FROM debit_note_items WHERE debit_note_id=%s", (dnid,))
        db.execute("DELETE FROM debit_notes WHERE id=%s", (dnid,))
        db.commit()
    flash('Debit note deleted', 'success')
    return redirect(url_for('debit_notes'))


# ── Stock Movement History ────────────────────────────────────────────────────

@app.route('/products/<int:pid>/movements')
def stock_movements(pid):
    with get_db() as db:
        product = db.execute("SELECT * FROM products WHERE id=%s", (pid,)).fetchone()
        warehouses = db.execute("SELECT * FROM warehouses").fetchall()
        wh_filter = request.args.get('wh', type=int)

        wh_cond = f"AND t.warehouse_id={wh_filter}" if wh_filter else ""
        wh_cond_from = f"AND t.from_wh_id={wh_filter}" if wh_filter else ""
        wh_cond_to = f"AND t.to_wh_id={wh_filter}" if wh_filter else ""

        movements = db.execute(f"""
            SELECT 'Purchase' as doc_type, pur.num as doc_ref, pur.doc_date,
                   pi.warehouse_id, wh.name as wh_name, pi.qty as qty_change, pur.supplier as counterpart
            FROM purchase_items pi
            JOIN purchases pur ON pur.id=pi.purchase_id
            JOIN warehouses wh ON wh.id=pi.warehouse_id
            WHERE pi.product_id=%s {wh_cond.replace('t.','pi.')}

            UNION ALL

            SELECT 'Sale', s.num, s.doc_date,
                   si.warehouse_id, wh.name, -si.qty, s.customer
            FROM sale_items si
            JOIN sales s ON s.id=si.sale_id
            JOIN warehouses wh ON wh.id=si.warehouse_id
            WHERE si.product_id=%s AND s.status NOT IN ('cancelled') {wh_cond.replace('t.','si.')}

            UNION ALL

            SELECT 'Write-off', wo.num, wo.doc_date,
                   wo.warehouse_id, wh.name, -wi.qty, wo.reason
            FROM writeoff_items wi
            JOIN writeoffs wo ON wo.id=wi.writeoff_id
            JOIN warehouses wh ON wh.id=wo.warehouse_id
            WHERE wi.product_id=%s {wh_cond.replace('t.','wo.')}

            UNION ALL

            SELECT 'Transfer (out)', 'TRF-'||CAST(t.id AS TEXT), t.doc_date,
                   t.from_wh_id, wh.name, -t.qty, 'To: '||(SELECT name FROM warehouses WHERE id=t.to_wh_id)
            FROM transfers t
            JOIN warehouses wh ON wh.id=t.from_wh_id
            WHERE t.product_id=%s {wh_cond_from.replace('t.from_wh_id','t.from_wh_id')}

            UNION ALL

            SELECT 'Transfer (in)', 'TRF-'||CAST(t.id AS TEXT), t.doc_date,
                   t.to_wh_id, wh.name, t.qty, 'From: '||(SELECT name FROM warehouses WHERE id=t.from_wh_id)
            FROM transfers t
            JOIN warehouses wh ON wh.id=t.to_wh_id
            WHERE t.product_id=%s {wh_cond_to.replace('t.to_wh_id','t.to_wh_id')}

            UNION ALL

            SELECT 'Credit Note', cn.num, cn.doc_date,
                   cn.warehouse_id, wh.name, cni.qty, 'Return from: '||cn.customer
            FROM credit_note_items cni
            JOIN credit_notes cn ON cn.id=cni.credit_note_id
            JOIN warehouses wh ON wh.id=cn.warehouse_id
            WHERE cni.product_id=%s AND cn.status='applied' {wh_cond.replace('t.','cn.')}

            UNION ALL

            SELECT 'Debit Note', dn.num, dn.doc_date,
                   dn.warehouse_id, wh.name, -dni.qty, 'Return to: '||dn.supplier
            FROM debit_note_items dni
            JOIN debit_notes dn ON dn.id=dni.debit_note_id
            JOIN warehouses wh ON wh.id=dn.warehouse_id
            WHERE dni.product_id=%s AND dn.status='applied' {wh_cond.replace('t.','dn.')}

            ORDER BY doc_date DESC, doc_ref DESC
        """, [pid]*7).fetchall()

        current_stock = {r['warehouse_id']: r['qty'] for r in
                         db.execute("SELECT warehouse_id, qty FROM stock WHERE product_id=%s", (pid,)).fetchall()}
    return render_template('stock_movements.html',
        product=product, movements=movements, warehouses=warehouses,
        wh_filter=wh_filter, current_stock=current_stock)


# ── Invoice Payments ──────────────────────────────────────────────────────────

@app.route('/sales/<int:sid>/payment', methods=['POST'])
@admin_required
def add_payment(sid):
    amount = float(request.form.get('amount', 0))
    if amount <= 0:
        flash('Amount must be positive', 'error')
        return redirect(url_for('view_sale', sid=sid))
    with get_db() as db:
        db.execute(
            "INSERT INTO invoice_payments(sale_id,payment_date,amount,method,notes,created_at) VALUES(%s,%s,%s,%s,%s,%s)",
            (sid, request.form['payment_date'], amount,
             request.form.get('method', 'bank'),
             request.form.get('notes', ''),
             datetime.utcnow().isoformat()))
        paid_total = db.execute("SELECT COALESCE(SUM(amount),0) AS s FROM invoice_payments WHERE sale_id=%s", (sid,)).fetchone()['s'] + amount
        sale_total = db.execute("SELECT total AS t FROM sales WHERE id=%s", (sid,)).fetchone()['t']
        db.execute("UPDATE sales SET paid=%s WHERE id=%s", (1 if paid_total >= sale_total else 0, sid))
        db.commit()
    flash('Payment recorded', 'success')
    return redirect(url_for('view_sale', sid=sid))


@app.route('/sales/<int:sid>/payment/<int:pid>/delete', methods=['POST'])
@admin_required
def delete_payment(sid, pid):
    with get_db() as db:
        db.execute("DELETE FROM invoice_payments WHERE id=%s AND sale_id=%s", (pid, sid))
        paid_total = db.execute("SELECT COALESCE(SUM(amount),0) AS s FROM invoice_payments WHERE sale_id=%s", (sid,)).fetchone()['s']
        sale_total = db.execute("SELECT total AS t FROM sales WHERE id=%s", (sid,)).fetchone()['t']
        db.execute("UPDATE sales SET paid=%s WHERE id=%s", (1 if paid_total >= sale_total else 0, sid))
        db.commit()
    flash('Payment removed', 'success')
    return redirect(url_for('view_sale', sid=sid))


# ── Export helpers ────────────────────────────────────────────────────────────

def generate_invoice_pdf(sale, items, customer=None, company=None, doc_title='INVOICE'):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    ACCENT   = colors.HexColor('#1a1a2e')
    ACCENT2  = colors.HexColor('#e8e6df')
    MUTED    = colors.HexColor('#666666')
    WHITE    = colors.white
    LIGHT_BG = colors.HexColor('#f8f8f6')

    buf = io.BytesIO()
    W, H = A4
    pdf = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=14*mm, bottomMargin=20*mm,
                            leftMargin=18*mm, rightMargin=18*mm)

    CURRENCY_SYM = {'GBP': '£', 'USD': '$', 'EUR': '€', 'PLN': 'zł'}
    sym = CURRENCY_SYM.get(sale.get('currency', 'GBP'), '£')

    styles = getSampleStyleSheet()
    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    normal    = ps('n', fontSize=9, leading=13)
    normal_r  = ps('nr', fontSize=9, leading=13, alignment=TA_RIGHT)
    small_m   = ps('sm', fontSize=8, leading=11, textColor=MUTED)
    small_mr  = ps('smr', fontSize=8, leading=11, textColor=MUTED, alignment=TA_RIGHT)
    label_s   = ps('lb', fontSize=7, leading=10, textColor=MUTED, fontName='Helvetica-Bold',
                   spaceAfter=1)
    inv_title = ps('it', fontSize=28, fontName='Helvetica-Bold', textColor=ACCENT, alignment=TA_RIGHT)
    co_name_s = ps('cn', fontSize=11, fontName='Helvetica-Bold', textColor=ACCENT, leading=14)
    bill_name = ps('bn', fontSize=9, fontName='Helvetica-Bold', leading=13)

    co = company or {}
    elements = []

    # ── TOP BANNER ────────────────────────────────────────────────────────────
    logo_cell = ''
    logo_key = co.get('co_logo', '')
    if logo_key:
        logo_path = os.path.normpath(os.path.join(UPLOAD_FOLDER, '..', logo_key))
        if os.path.exists(logo_path):
            try:
                img = Image(logo_path)
                img._restrictSize(55*mm, 22*mm)
                logo_cell = img
            except Exception:
                pass

    banner = Table(
        [[logo_cell, Paragraph(doc_title, inv_title)]],
        colWidths=[90*mm, 80*mm]
    )
    banner.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(banner)
    elements.append(Spacer(1, 3*mm))
    elements.append(HRFlowable(width='100%', thickness=2, color=ACCENT))
    elements.append(Spacer(1, 5*mm))

    # ── FROM / INVOICE META ───────────────────────────────────────────────────
    co_block = []
    if co.get('co_company'):
        co_block.append(Paragraph(co['co_company'], co_name_s))
    for line in filter(None, [co.get('co_address'), co.get('co_address2')]):
        co_block.append(Paragraph(line, small_m))
    addr_parts = ' '.join(filter(None, [co.get('co_city',''), co.get('co_postcode','')]))
    if addr_parts: co_block.append(Paragraph(addr_parts, small_m))
    if co.get('co_country'): co_block.append(Paragraph(co['co_country'], small_m))
    if co.get('co_vat'):     co_block.append(Paragraph(f"VAT No: {co['co_vat']}", small_m))

    paid_badge = 'PAID' if sale.get('paid') else 'UNPAID'
    paid_color = '#1a8a4a' if sale.get('paid') else '#c0392b'

    meta_table = Table([
        [Paragraph('<font color="#888888" size="7">INVOICE NUMBER</font>', label_s),
         Paragraph(f"<b>{sale['num']}</b>", normal_r)],
        [Paragraph('<font color="#888888" size="7">DATE</font>', label_s),
         Paragraph(sale['doc_date'], normal_r)],
        [Paragraph('<font color="#888888" size="7">STATUS</font>', label_s),
         Paragraph(f"<font color='{paid_color}'><b>{paid_badge}</b></font>", normal_r)],
    ], colWidths=[30*mm, 40*mm])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),
    ]))

    from_col = [[Paragraph('<font color="#888888" size="7">FROM</font>', label_s)] + co_block]

    info_table = Table(
        [[from_col[0], meta_table]],
        colWidths=[100*mm, 70*mm]
    )
    info_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 6*mm))

    # ── BILL TO ───────────────────────────────────────────────────────────────
    elements.append(HRFlowable(width='100%', thickness=0.5, color=ACCENT2))
    elements.append(Spacer(1, 4*mm))

    bill_block = [Paragraph('<font color="#888888" size="7">BILL TO</font>', label_s)]
    if customer:
        if customer.get('company'):
            bill_block.append(Paragraph(customer['company'], bill_name))
        for line in filter(None, [customer.get('address'), customer.get('address2')]):
            bill_block.append(Paragraph(line, small_m))
        if customer.get('city'):     bill_block.append(Paragraph(customer['city'], small_m))
        if customer.get('postcode'): bill_block.append(Paragraph(customer['postcode'], small_m))
        if customer.get('country'):    bill_block.append(Paragraph(customer['country'], small_m))
        if customer.get('vat_number'): bill_block.append(Paragraph(f"VAT No: {customer['vat_number']}", small_m))

    elements.append(Table([[bill_block]], colWidths=[170*mm]))
    elements.append(Spacer(1, 6*mm))

    # ── LINE ITEMS TABLE ──────────────────────────────────────────────────────
    has_disc = any(i.get('discount_pct', 0) for i in items)
    has_ctns = any((i.get('carton_qty') or 0) > 0 for i in items)

    hdr = ['Description', 'Qty', 'Unit']
    col_w_map = [80*mm, 14*mm, 12*mm]
    if has_ctns:
        hdr.append('Pcs/CTN')
        col_w_map.append(16*mm)
        hdr.append('CTNs')
        col_w_map.append(14*mm)
    hdr.append('Unit Price')
    col_w_map.append(26*mm)
    if has_disc:
        hdr.append('Disc%')
        col_w_map.append(14*mm)
    hdr.append('Amount')
    col_w_map.append(26*mm)
    # trim description col to fit page
    used = sum(col_w_map[1:])
    col_w_map[0] = 170*mm - used
    col_w = col_w_map
    span_start = len(hdr) - 2

    data = [hdr]
    for i in items:
        ctn_qty = i.get('carton_qty') or 0
        line_total = i['qty'] * i['price'] * (1 - i.get('discount_pct', 0) / 100)
        row = [i['product_name'], f"{i['qty']:g}", i.get('unit', '')]
        if has_ctns:
            row.append(f"{int(ctn_qty)}" if ctn_qty > 0 else '—')
            row.append(f"{i['qty']/ctn_qty:g}" if ctn_qty > 0 else '—')
        row.append(f"{sym}{i['price']:,.4f}")
        if has_disc:
            row.append(f"{i.get('discount_pct',0):g}%")
        row.append(f"{sym}{line_total:,.2f}")
        data.append(row)

    n = len(items)
    num_cols = len(hdr)

    def total_row(label, value, bold=False):
        fn = 'Helvetica-Bold' if bold else 'Helvetica'
        sz = 10 if bold else 9
        empty = [''] * span_start
        return empty + [
            Paragraph(f"<font name='{fn}' size='{sz}'>{label}</font>", normal_r),
            Paragraph(f"<font name='{fn}' size='{sz}'>{value}</font>", normal_r),
        ]

    line_sum = sum(i['qty'] * i['price'] * (1 - i.get('discount_pct', 0) / 100) for i in items)
    data.append(total_row('Subtotal', f"{sym}{line_sum:,.2f}"))
    if sale.get('discount', 0):
        disc_amt = line_sum * sale['discount'] / 100
        data.append(total_row(f"Discount ({sale['discount']:g}%)", f"−{sym}{disc_amt:,.2f}"))
    if sale.get('tax_pct', 0):
        data.append(total_row(f"VAT ({sale['tax_pct']:g}%)", f"{sym}{sale['tax_amount']:,.2f}"))
    data.append(total_row('TOTAL DUE', f"{sym}{sale['total']:,.2f}", bold=True))
    if has_ctns:
        total_c = sum(i['qty'] / i['carton_qty'] for i in items if (i.get('carton_qty') or 0) > 0)
        data.append(total_row('Total CTNs', f"{total_c:g}"))
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        # header row
        ('BACKGROUND', (0,0), (-1,0), ACCENT),
        ('TEXTCOLOR',  (0,0), (-1,0), WHITE),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 8.5),
        ('BOTTOMPADDING', (0,0), (-1,0), 7),
        ('TOPPADDING',    (0,0), (-1,0), 7),
        # item rows
        ('FONTSIZE',  (0,1), (-1,n), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,n), [WHITE, LIGHT_BG]),
        ('LINEBELOW', (0,1), (-1,n), 0.3, ACCENT2),
        ('TOPPADDING',    (0,1), (-1,n), 6),
        ('BOTTOMPADDING', (0,1), (-1,n), 6),
        # totals rows
        ('FONTSIZE',  (0,n+1), (-1,-1), 9),
        ('TOPPADDING',    (0,n+1), (-1,-1), 5),
        ('BOTTOMPADDING', (0,n+1), (-1,-1), 5),
        ('LINEABOVE', (0,n+1), (-1,n+1), 0.8, ACCENT2),
        ('SPAN',      (0,n+1), (span_start-1,-1)),
        ('LINEABOVE', (span_start,-1), (-1,-1), 1, ACCENT),
        ('TOPPADDING',    (0,-1), (-1,-1), 8),
        ('BOTTOMPADDING', (0,-1), (-1,-1), 8),
        # all
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('LEFTPADDING',  (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t)

    # ── NOTES ─────────────────────────────────────────────────────────────────
    if sale.get('notes'):
        elements.append(Spacer(1, 5*mm))
        elements.append(Paragraph(f"<b>Notes:</b> {sale['notes']}", small_m))

    # ── FOOTER: bank details ──────────────────────────────────────────────────
    bank_parts = []
    if co.get('co_bank_name'):      bank_parts.append(('Bank', co['co_bank_name']))
    if co.get('co_sort_code'):      bank_parts.append(('Sort Code', co['co_sort_code']))
    if co.get('co_account_number'): bank_parts.append(('Account No', co['co_account_number']))
    if bank_parts:
        elements.append(Spacer(1, 8*mm))
        elements.append(HRFlowable(width='100%', thickness=0.5, color=ACCENT2))
        elements.append(Spacer(1, 3*mm))
        bank_cells = [[Paragraph(f"<font color='#888888' size='7'>{k}</font><br/><b>{v}</b>", normal)
                       for k, v in bank_parts]]
        bw = 170*mm / max(len(bank_parts), 1)
        bank_table = Table(bank_cells, colWidths=[bw]*len(bank_parts))
        bank_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
        ]))
        elements.append(bank_table)

    pdf.build(elements)
    return buf.getvalue()


def generate_pdf(title, doc, items, doc_type='purchase'):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    buf = io.BytesIO()
    pdf = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=20*mm, rightMargin=20*mm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 6*mm)]
    meta = [('Customer' if doc_type=='sale' else 'Supplier', doc.get('customer' if doc_type=='sale' else 'supplier','')),
            ('Date', doc.get('doc_date','')), ('Currency', doc.get('currency','GBP')), ('Status', doc.get('status',''))]
    for k, v in meta:
        elements.append(Paragraph(f"<b>{k}:</b> {v}", styles['Normal']))
    elements.append(Spacer(1, 6*mm))
    if doc_type == 'sale':
        hdr = ['Product', 'Qty', 'Unit', 'Price', 'Disc%', 'Subtotal']
        data = [hdr] + [[i['product_name'], f"{i['qty']:g}", i.get('unit',''),
                         f"£{i['price']:,.2f}", f"{i.get('discount_pct',0)}%",
                         f"£{i['qty']*i['price']*(1-i.get('discount_pct',0)/100):,.2f}"] for i in items]
    else:
        hdr = ['Product', 'Qty', 'Unit', 'Unit Cost', 'Subtotal']
        data = [hdr] + [[i['product_name'], f"{i['qty']:g}", i.get('unit',''),
                         f"£{i['price']:,.2f}", f"£{i['qty']*i['price']:,.2f}"] for i in items]
    data.append(['', '', '', 'TOTAL', f"£{doc['total']:,.2f}"])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a6ef5')), ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('FONTSIZE',(0,0),(-1,-1),9),
        ('ROWBACKGROUNDS',(0,1),(-1,-2),[colors.white,colors.HexColor('#f5f4f0')]),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#e8e6df')),
        ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'), ('ALIGN',(1,0),(-1,-1),'RIGHT'), ('PADDING',(0,0),(-1,-1),6),
    ]))
    elements.append(t)
    pdf.build(elements)
    return buf.getvalue()


def generate_excel(title, doc, items, doc_type='purchase', contact=None, company=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, GradientFill
    from openpyxl.utils import get_column_letter

    CURRENCY_SYM = {'GBP': '£', 'USD': '$', 'EUR': '€', 'PLN': 'zł'}
    sym = CURRENCY_SYM.get(doc.get('currency', 'GBP'), '£')
    co  = company or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # ── Styles ────────────────────────────────────────────────────────────────
    DARK   = '1A1A2E'
    LIGHT  = 'F8F8F6'
    ALT    = 'F0EDE8'
    MUTED  = '888888'
    RED    = 'C0392B'
    GREEN  = '1A8A4A'

    def fill(hex_): return PatternFill('solid', fgColor=hex_)
    def border():
        s = Side(style='thin', color='E8E6DF')
        return Border(left=s, right=s, top=s, bottom=s)
    def bottom_border(color='E8E6DF', thick=False):
        s = Side(style='medium' if thick else 'thin', color=color)
        return Border(bottom=s)

    def cell(row, col, value='', bold=False, size=10, color=None, bg=None, align='left', italic=False, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=size, color=color or '000000', italic=italic)
        if bg:   c.fill = fill(bg)
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        if num_fmt: c.number_format = num_fmt
        return c

    # ── Doc type label ────────────────────────────────────────────────────────
    doc_label = 'INVOICE' if doc_type == 'sale' else 'PURCHASE ORDER'
    is_sale   = doc_type == 'sale'

    # ── ROW 1 & 2 will be written AFTER we know num_cols ─────────────────────
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 6

    # ── ROW 3-7: FROM (left) + Meta (right) ───────────────────────────────────
    from_lines = []
    for line in filter(None, [co.get('co_address'), co.get('co_address2')]):
        from_lines.append(line)
    addr = ' '.join(filter(None, [co.get('co_city',''), co.get('co_postcode','')]))
    if addr: from_lines.append(addr)
    if co.get('co_country'):  from_lines.append(co['co_country'])
    if co.get('co_vat'):      from_lines.append(f"VAT: {co['co_vat']}")

    meta_rows = [
        ('NUMBER',   doc.get('num', title)),
        ('DATE',     doc.get('doc_date', '')),
    ]
    if is_sale:
        paid_badge = 'PAID' if doc.get('paid') else 'UNPAID'
        meta_rows.append(('STATUS', paid_badge))
    else:
        meta_rows.append(('WAREHOUSE', doc.get('wh_name', '')))
    meta_rows.append(('CURRENCY', doc.get('currency', 'GBP')))

    info_start = 3
    for i, line in enumerate(from_lines[:5]):
        r = info_start + i
        ws.row_dimensions[r].height = 16
        cell(r, 1, line, size=9, color=MUTED)

    for i, (lbl, val) in enumerate(meta_rows):
        r = info_start + i
        ws.row_dimensions[r].height = 16
        # label in col 6, value in col 7 (will adjust after col count known)
        cell(r, 6, lbl, size=7, color=MUTED, bold=True)

    # ── ROW: blank ────────────────────────────────────────────────────────────
    blank_row = info_start + max(len(from_lines[:5]), len(meta_rows)) + 1
    ws.row_dimensions[blank_row].height = 4

    # ── BILL TO / SUPPLIER section ────────────────────────────────────────────
    bt_row = blank_row + 1
    contact_label = 'BILL TO' if is_sale else 'SUPPLIER'
    ws.row_dimensions[bt_row].height = 14
    cell(bt_row, 1, contact_label, size=7, color=MUTED, bold=True)
    ct = contact or {}
    contact_name = ct.get('company') or ct.get('name') or doc.get('customer' if is_sale else 'supplier', '')
    cell(bt_row + 1, 1, contact_name, bold=True, size=10)
    ws.row_dimensions[bt_row + 1].height = 18
    ct_lines = []
    for line in filter(None, [ct.get('address'), ct.get('address2')]):
        ct_lines.append(line)
    city_post = ' '.join(filter(None, [ct.get('city',''), ct.get('postcode','')]))
    if city_post: ct_lines.append(city_post)
    if ct.get('country'):    ct_lines.append(ct['country'])
    if ct.get('vat_number'): ct_lines.append(f"VAT: {ct['vat_number']}")
    for i, line in enumerate(ct_lines):
        r = bt_row + 2 + i
        ws.row_dimensions[r].height = 15
        cell(r, 1, line, size=9, color=MUTED)

    # ── LINE ITEMS TABLE ──────────────────────────────────────────────────────
    items_start = bt_row + 2 + len(ct_lines) + 2

    has_disc = any(item.get('discount_pct', 0) for item in items)
    has_ctns = any((item.get('carton_qty') or 0) > 0 for item in items)

    headers = ['#', 'Description', 'Qty', 'Unit']
    if has_ctns:
        headers += ['Pcs/CTN', 'CTNs']
    headers.append('Unit Price')
    if has_disc:
        headers.append('Disc%')
    headers.append('Amount')
    num_cols = len(headers)

    # Now write banner row (row 1) with dark background across all cols
    for c in range(1, num_cols + 1):
        bc = ws.cell(1, c)
        bc.fill      = fill(DARK)
        bc.font      = Font(size=14, color='FFFFFF')
        bc.alignment = Alignment(vertical='center')
        ws.cell(2, c).fill = fill(DARK)
    # Company name left, doc label right
    ws.cell(1, 1).value = co.get('co_company', 'NEON')
    ws.cell(1, 1).font  = Font(bold=True, size=14, color='FFFFFF')
    ws.cell(1, 1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(1, num_cols).value = doc_label
    ws.cell(1, num_cols).font  = Font(bold=True, size=14, color='FFFFFF')
    ws.cell(1, num_cols).alignment = Alignment(horizontal='right', vertical='center')

    # Meta right-side values (col 6 = label, col 7+ = value, capped to num_cols)
    meta_label_col = max(num_cols - 2, 4)
    meta_val_col   = num_cols
    for i, (lbl, val) in enumerate(meta_rows):
        r = info_start + i
        # clear old label cell (col 6 may differ)
        ws.cell(r, 6).value = None
        c_lbl = ws.cell(r, meta_label_col)
        c_lbl.value = lbl
        c_lbl.font  = Font(size=7, bold=True, color=MUTED)
        c_lbl.alignment = Alignment(horizontal='right')
        c_val = ws.cell(r, meta_val_col)
        c_val.value = val
        c_val.font  = Font(size=9, bold=True)
        c_val.alignment = Alignment(horizontal='right')
        if lbl == 'STATUS':
            c_val.font = Font(size=9, bold=True, color=GREEN if val == 'PAID' else RED)

    # Header row
    ws.row_dimensions[items_start].height = 22
    for col, h in enumerate(headers, 1):
        c = ws.cell(items_start, col, h)
        c.font      = Font(bold=True, size=9, color='FFFFFF')
        c.fill      = fill(DARK)
        c.alignment = Alignment(horizontal='right' if col > 2 else 'left', vertical='center')
        c.border    = border()

    # Item rows
    for idx, item in enumerate(items):
        r = items_start + 1 + idx
        ws.row_dimensions[r].height = 18
        row_bg = None if idx % 2 == 0 else ALT
        ctn_qty = item.get('carton_qty') or 0
        d       = item.get('discount_pct', 0) or 0
        line_total = item['qty'] * item['price'] * (1 - d / 100)

        row_vals = [idx + 1, item['product_name'], item['qty'], item.get('unit', '')]
        if has_ctns:
            row_vals.append(int(ctn_qty) if ctn_qty > 0 else '—')
            row_vals.append(round(item['qty'] / ctn_qty, 2) if ctn_qty > 0 else '—')
        row_vals.append(item['price'])
        if has_disc:
            row_vals.append(d)
        row_vals.append(line_total)

        for col, val in enumerate(row_vals, 1):
            c = ws.cell(r, col, val)
            c.font      = Font(size=9)
            c.alignment = Alignment(horizontal='right' if col > 2 else 'left', vertical='center')
            c.border    = border()
            if row_bg:  c.fill = fill(row_bg)
            h_name = headers[col - 1]
            if h_name == 'Unit Price':
                c.number_format = f'"{sym}"#,##0.0000'
            elif h_name == 'Amount':
                c.number_format = f'"{sym}"#,##0.00'
            elif h_name in ('Qty', 'CTNs'):
                c.number_format = '#,##0.###'
            elif h_name == 'Disc%':
                c.number_format = '0.##"%"'

    # ── Totals ────────────────────────────────────────────────────────────────
    def total_row(r, label, value, bold=False, color=None):
        ws.row_dimensions[r].height = 18
        lc = ws.cell(r, num_cols - 1, label)
        vc = ws.cell(r, num_cols, value)
        lc.font = vc.font = Font(bold=bold, size=9, color=color or '000000')
        lc.alignment = Alignment(horizontal='right', vertical='center')
        vc.alignment = Alignment(horizontal='right', vertical='center')
        vc.number_format = f'"{sym}"#,##0.00'
        lc.border = bottom_border()
        vc.border = bottom_border()

    line_sum = sum(i['qty'] * i['price'] * (1 - (i.get('discount_pct') or 0) / 100) for i in items)
    t_row = items_start + len(items) + 2
    ws.row_dimensions[t_row - 1].height = 6  # gap
    total_row(t_row,     'Subtotal',      line_sum)
    t_row += 1
    if doc.get('discount', 0):
        disc_amt = line_sum * doc['discount'] / 100
        total_row(t_row, f"Discount ({doc['discount']:g}%)", -disc_amt, color=RED)
        t_row += 1
    if doc.get('tax_pct', 0):
        total_row(t_row, f"VAT ({doc['tax_pct']:g}%)", doc.get('tax_amount', 0))
        t_row += 1
    # TOTAL row with thick top border
    ws.row_dimensions[t_row].height = 22
    lc = ws.cell(t_row, num_cols - 1, 'TOTAL')
    vc = ws.cell(t_row, num_cols, doc.get('total', 0))
    lc.font = vc.font = Font(bold=True, size=11)
    lc.alignment = Alignment(horizontal='right', vertical='center')
    vc.alignment = Alignment(horizontal='right', vertical='center')
    vc.number_format = f'"{sym}"#,##0.00'
    thick = Side(style='medium', color=DARK)
    lc.border = Border(top=thick, bottom=thick)
    vc.border = Border(top=thick, bottom=thick)
    t_row += 1

    # CTN total
    if has_ctns:
        total_ctns = sum(i['qty'] / i['carton_qty'] for i in items if (i.get('carton_qty') or 0) > 0)
        ws.row_dimensions[t_row].height = 16
        lc = ws.cell(t_row, num_cols - 1, 'Total CTNs')
        vc = ws.cell(t_row, num_cols, round(total_ctns, 2))
        lc.font = vc.font = Font(size=9, color=MUTED)
        lc.alignment = Alignment(horizontal='right'); vc.alignment = Alignment(horizontal='right')
        t_row += 1

    # ── Notes ─────────────────────────────────────────────────────────────────
    if doc.get('notes'):
        t_row += 1
        ws.row_dimensions[t_row].height = 16
        cell(t_row, 1, 'Notes:', bold=True, size=9, color=MUTED)
        cell(t_row, 2, doc['notes'], size=9, color=MUTED)
        t_row += 1

    # ── Bank details ──────────────────────────────────────────────────────────
    bank_parts = []
    if co.get('co_bank_name'):      bank_parts.append(('Bank', co['co_bank_name']))
    if co.get('co_sort_code'):      bank_parts.append(('Sort Code', co['co_sort_code']))
    if co.get('co_account_number'): bank_parts.append(('Account No', co['co_account_number']))
    if bank_parts:
        t_row += 1
        for col_i, (lbl, val) in enumerate(bank_parts):
            ws.row_dimensions[t_row].height = 14
            ws.row_dimensions[t_row + 1].height = 16
            c_lbl = ws.cell(t_row,     col_i * 2 + 1, lbl)
            c_val = ws.cell(t_row + 1, col_i * 2 + 1, val)
            c_lbl.font = Font(size=7, bold=True, color=MUTED)
            c_val.font = Font(size=9, bold=True)

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        1: 4,   # #
        2: 36,  # Description
        3: 8,   # Qty
        4: 7,   # Unit
    }
    extra_col = 5
    if has_ctns:
        col_widths[extra_col] = 9   # Pcs/CTN
        col_widths[extra_col + 1] = 8  # CTNs
        extra_col += 2
    col_widths[extra_col] = 13   # Unit Price
    if has_disc:
        col_widths[extra_col + 1] = 7
        col_widths[extra_col + 2] = 13
    else:
        col_widths[extra_col + 1] = 13  # Amount

    for col_i in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = col_widths.get(col_i, 12)

    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Global Search ─────────────────────────────────────────────────────────────

@app.route('/search')
def global_search():
    q = request.args.get('q', '').strip()
    results = {}
    if q:
        like = f'%{q}%'
        with get_db() as db:
            results['products'] = db.execute(
                "SELECT id, sku, name, category, unit FROM products "
                "WHERE name LIKE %s OR sku LIKE %s OR barcode LIKE %s ORDER BY name LIMIT 20",
                (like, like, like)
            ).fetchall()

            results['contacts'] = db.execute(
                "SELECT id, name, type, email, phone FROM contacts "
                "WHERE name LIKE %s OR email LIKE %s OR phone LIKE %s ORDER BY name LIMIT 20",
                (like, like, like)
            ).fetchall()

            results['sales'] = db.execute(
                "SELECT s.id, s.num, s.doc_date, s.status, s.total, c.name AS contact_name "
                "FROM sales s LEFT JOIN contacts c ON s.customer_id=c.id "
                "WHERE s.num LIKE %s OR s.customer LIKE %s OR c.name LIKE %s ORDER BY s.doc_date DESC LIMIT 20",
                (like, like, like)
            ).fetchall()

            results['purchases'] = db.execute(
                "SELECT p.id, p.num, p.doc_date, p.status, p.total, c.name AS contact_name "
                "FROM purchases p LEFT JOIN contacts c ON p.supplier_id=c.id "
                "WHERE p.num LIKE %s OR p.supplier LIKE %s OR c.name LIKE %s ORDER BY p.doc_date DESC LIMIT 20",
                (like, like, like)
            ).fetchall()

            results['quotes'] = db.execute(
                "SELECT q.id, q.num, q.doc_date, q.status, q.total, c.name AS contact_name "
                "FROM quotes q LEFT JOIN contacts c ON q.customer_id=c.id "
                "WHERE q.num LIKE %s OR q.customer LIKE %s OR c.name LIKE %s ORDER BY q.doc_date DESC LIMIT 20",
                (like, like, like)
            ).fetchall()

            results['writeoffs'] = db.execute(
                "SELECT w.id, w.num, w.doc_date, w.reason, wh.name AS wh_name "
                "FROM writeoffs w LEFT JOIN warehouses wh ON w.warehouse_id=wh.id "
                "WHERE w.num LIKE %s OR w.reason LIKE %s ORDER BY w.doc_date DESC LIMIT 20",
                (like, like)
            ).fetchall()

            results['credit_notes'] = db.execute(
                "SELECT cn.id, cn.num, cn.doc_date, cn.status, cn.total, c.name AS contact_name "
                "FROM credit_notes cn LEFT JOIN contacts c ON cn.customer_id=c.id "
                "WHERE cn.num LIKE %s OR cn.customer LIKE %s OR c.name LIKE %s ORDER BY cn.doc_date DESC LIMIT 20",
                (like, like, like)
            ).fetchall()

            results['debit_notes'] = db.execute(
                "SELECT dn.id, dn.num, dn.doc_date, dn.status, dn.total, c.name AS supplier "
                "FROM debit_notes dn LEFT JOIN contacts c ON dn.supplier_id=c.id "
                "WHERE dn.num LIKE %s OR c.name LIKE %s ORDER BY dn.doc_date DESC LIMIT 20",
                (like, like)
            ).fetchall()

            results['transactions'] = db.execute(
                'SELECT t.id, t.doc_date, t.type, t.amount, t."desc" AS notes, c.name AS contact_name '
                'FROM transactions t LEFT JOIN contacts c ON t.contact_id=c.id '
                'WHERE t.type LIKE %s OR t."desc" LIKE %s OR c.name LIKE %s ORDER BY t.doc_date DESC LIMIT 20',
                (like, like, like)
            ).fetchall()

            results['warehouses'] = db.execute(
                "SELECT id, name, location FROM warehouses WHERE name LIKE %s OR location LIKE %s LIMIT 10",
                (like, like)
            ).fetchall()

    total = sum(len(v) for v in results.values())
    return render_template('search.html', q=q, results=results, total=total)


# ── API ───────────────────────────────────────────────────────────────────────

@app.route('/api/stock/<int:pid>/<int:wid>')
def api_stock(pid, wid):
    with get_db() as db:
        row = db.execute("SELECT qty FROM stock WHERE product_id=%s AND warehouse_id=%s", (pid, wid)).fetchone()
    return jsonify({'qty': row['qty'] if row else 0})


@app.route('/api/product/<int:pid>')
def api_product(pid):
    with get_db() as db:
        p = db.execute("SELECT * FROM products WHERE id=%s", (pid,)).fetchone()
        if not p: return jsonify({})
        stock = {str(r['warehouse_id']): r['qty'] for r in db.execute("SELECT warehouse_id,qty FROM stock WHERE product_id=%s", (pid,)).fetchall()}
    return jsonify({**dict(p), 'stock': stock})


@app.route('/api/contact/<int:cid>')
def api_customer(cid):
    with get_db() as db:
        c = db.execute("SELECT * FROM contacts WHERE id=%s", (cid,)).fetchone()
    return jsonify(dict(c) if c else {})


init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True, port=5001)
